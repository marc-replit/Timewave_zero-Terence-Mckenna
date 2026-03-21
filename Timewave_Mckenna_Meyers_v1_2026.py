# =============================================================================
# TIMEWAVE ZERO v3.12
# Run: streamlit run Timewave_Mckenna_Meyers_v1_2026.py
#
# NEW IN v3.11 (from v3.10):
# 1.  FIX [Critical]: st.set_page_config() moved to first Streamlit call (was
#     after session_state init — violates Streamlit's API requirements).
# 2.  FIX [Critical/Math]: Python code snippet (Math tab C) used a wrong fake
#     KELLEY_64 array (64 values × 6) that does not match the app's actual
#     Kelley primitive. Replaced with correct full KELLEY_384 (384 values).
#     Downloaded timewave.py now produces results identical to the app.
# 3.  FIX [Animation]: Build Animation used st.session_state.levels (stale)
#     instead of the local `levels` variable (current). Now uses `levels`.
# 4.  FIX [Animation]: Added stale-data warning when primitive / hexagram /
#     levels change after a Build without rebuilding the animation.
# 5.  FIX [Manual]: "All four primitives" corrected to "All five primitives"
#     in the Comparison tab description.
#
# CARRIED FROM v3.10:
# 1.  FIX: anim_resolution slider conflict — removed value= (key= alone is correct)
# 2.  FIX: Green banner now persists until next Process/Throw (not cleared on rerender)
# 3.  FIX: Comparison tab resonance markers now show all 5 primitives with own colours
# 4.  FIX: Animation tab flip on Build — removed st.spinner() (streams UI = flips tab)
#
# CARRIED FROM v3.7:
# 1.  New primitive "Original McKenna" added — raw first-order differences
# 2.  New tab [Original McKenna] after [HuangTi], fully computed like others
# 3.  Comparison + Resonance tabs updated to include Original McKenna
# 4.  HEXAGRAM_PATTERNS replaced with corrected classical King Wen patterns
# 5.  New tab [Math] after [Manual] with full mathematical documentation
# 6.  Tab indices shifted: Comparison=5, Resonance=6, Animation=7, Hexagrams=8, Manual=9, Math=10
# =============================================================================

import streamlit as st
import plotly.graph_objects as go
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from scipy.signal import argrelextrema
import random
import calendar
import pandas as pd

# =============================================================================
# PAGE CONFIG — must be the FIRST Streamlit call in the script
# =============================================================================

st.set_page_config(
    page_title="Novelty Explorer",
    page_icon="☯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# CONFIGURATION & DATA
# =============================================================================

APP_VERSION = "3.12.0"
COLORS_PRIMITIVES = {
    'Kelley':           '#4C72B0',
    'Watkins':          '#DD8452',
    'Sheliak':          '#55A868',
    'HuangTi':          '#C44E52',
    'Original McKenna': '#9B59B6'
}

PRIMITIVES = {
    "Kelley": np.array([
        50,48,48,40,41,38,31,31,31,22,25,28,25,59,44,62,54,48,38,22,20,20,22,23,
        27,36,40,48,72,62,63,52,49,41,43,40,39,42,44,46,42,41,38,37,37,39,37,41,
        45,41,44,44,42,47,59,69,69,60,63,60,42,42,37,40,49,47,41,37,44,34,29,29,
        43,54,49,37,33,43,46,39,37,42,47,43,37,37,39,42,40,41,35,38,40,41,37,38,
        38,33,35,45,50,55,35,29,35,37,39,38,43,52,54,65,31,41,44,44,42,29,24,24,
        32,39,32,38,44,34,39,46,51,64,60,64,49,46,51,48,48,45,48,49,51,51,34,29,
        41,41,46,44,46,50,45,46,44,45,43,44,58,40,44,40,43,42,47,32,31,30,32,26,
        62,79,75,71,79,76,31,31,33,38,41,41,33,18,23,27,29,30,38,47,47,30,31,28,
        34,28,24,38,37,34,33,35,29,32,27,26,27,33,18,36,28,22,22,26,24,24,26,27,
        35,40,44,52,76,66,47,28,29,27,33,36,43,40,32,44,32,31,36,17,13,37,39,43,
        39,31,28,38,36,41,53,63,63,54,57,54,42,42,37,40,49,47,51,43,36,46,53,53,
        51,48,51,53,51,51,44,39,41,34,25,29,49,61,61,34,32,33,41,26,26,29,27,28,
        44,43,45,41,44,49,47,43,35,53,31,36,37,36,34,31,33,29,22,22,20,49,42,46,
        66,65,74,80,86,76,55,58,53,56,60,56,31,34,29,42,30,39,30,57,59,75,58,67,
        41,41,38,32,34,38,45,46,44,45,43,44,76,70,74,34,35,34,39,38,37,40,38,40,
        32,27,29,47,49,48,37,37,39,44,47,47,45,54,59,49,49,42,30,27,27,30,25,28
    ], dtype=float),
    
    "Watkins": np.array([
        50,48,44,42,17,30,31,31,31,22,25,28,25,59,44,62,54,48,38,22,20,20,22,23,
        27,36,40,48,72,62,63,52,49,41,19,16,29,50,36,54,30,33,46,45,47,31,29,33,
        37,33,36,36,34,39,67,77,77,52,55,52,34,34,29,32,41,39,33,45,52,42,37,37,
        19,82,77,65,61,71,74,35,33,18,23,19,29,29,31,50,48,49,43,30,48,41,45,46,
        30,41,43,37,42,47,27,21,27,29,31,30,51,60,62,57,23,33,36,36,34,21,16,16,
        24,31,40,46,52,42,47,54,59,72,68,72,41,38,43,40,40,37,40,41,43,59,42,37,
        33,33,38,36,38,42,37,38,36,37,35,36,66,32,36,32,35,34,47,24,23,22,24,18,
        70,87,83,79,87,84,23,23,25,30,33,33,41,26,31,19,21,22,46,55,55,22,23,20,
        26,36,32,30,29,26,25,43,37,40,19,18,19,41,26,44,36,30,30,34,32,32,34,35,
        27,48,52,60,84,74,39,20,21,19,41,44,51,32,24,36,24,23,44,25,21,45,47,51,
        31,39,36,30,28,33,61,71,71,46,49,46,34,34,29,32,41,39,59,51,44,54,61,61,
        43,40,43,45,43,43,36,31,33,42,33,37,41,69,69,26,24,25,49,18,18,21,19,20,
        36,35,37,33,36,41,39,51,43,45,23,28,29,28,26,23,25,21,30,30,28,41,34,38,
        74,73,82,88,94,84,47,66,61,48,52,48,23,42,37,34,22,31,38,65,67,83,50,59,
        33,33,30,24,26,30,37,38,36,37,35,36,84,62,66,26,27,26,31,30,29,32,30,48,
        40,35,37,39,41,40,29,29,31,36,39,39,53,62,67,41,41,34,22,19,19,22,17,20,
    ], dtype=float),
    
    "Sheliak": np.array([
        4,7,3,5,5,5,4,3,5,3,5,7,3,3,5,3,3,7,4,5,4,3,3,3,
        4,5,3,7,3,7,4,3,4,5,5,5,3,5,7,5,4,3,5,3,4,5,4,3,
        5,7,4,6,6,6,5,4,6,4,6,8,4,4,6,4,4,8,5,6,5,4,4,4,
        5,6,4,8,4,8,5,4,5,6,6,6,4,6,8,6,5,4,6,4,5,6,5,4,
        6,8,5,7,7,7,6,5,7,5,7,9,5,5,7,5,5,9,6,7,6,5,5,5,
        6,7,5,9,5,9,6,5,6,7,7,7,5,7,9,7,6,5,7,5,6,7,6,5,
        7,9,6,8,8,8,7,6,8,6,8,10,6,6,8,6,6,10,7,8,7,6,6,6,
        7,8,6,10,6,10,7,6,7,8,8,8,6,8,10,8,7,6,8,6,7,8,7,6,
        8,10,7,9,9,9,8,7,9,7,9,11,7,7,9,7,7,11,8,9,8,7,7,7,
        8,9,7,11,7,11,8,7,8,9,9,9,7,9,11,9,8,7,9,7,8,9,8,7,
        9,11,8,10,10,10,9,8,10,8,10,12,8,8,10,8,8,12,9,10,9,8,8,8,
        9,10,8,12,8,12,9,8,9,10,10,10,8,10,12,10,9,8,10,8,9,10,9,8,
        10,12,9,11,11,11,10,9,11,9,11,13,9,9,11,9,9,13,10,11,10,9,9,9,
        10,11,9,13,9,13,10,9,10,11,11,11,9,11,13,11,10,9,11,9,10,11,10,9,
        11,13,10,12,12,12,11,10,12,10,12,14,10,10,12,10,10,14,11,12,11,10,10,10,
        11,12,10,14,10,14,11,10,11,12,12,12,10,12,14,12,11,10,12,10,11,12,11,10
    ], dtype=float),
    "HuangTi": np.array([
        2,5,1,3,3,3,2,1,3,1,3,5,1,1,3,1,1,5,2,3,2,1,1,1,
        2,3,1,5,1,5,2,1,2,3,3,3,1,3,5,3,2,1,3,1,2,3,2,1,
        3,5,2,4,4,4,3,2,4,2,4,6,2,2,4,2,2,6,3,4,3,2,2,2,
        3,4,2,6,2,6,3,2,3,4,4,4,2,4,6,4,3,2,4,2,3,4,3,2,
        4,6,3,5,5,5,4,3,5,3,5,7,3,3,5,3,3,7,4,5,4,3,3,3,
        4,5,3,7,3,7,4,3,4,5,5,5,3,5,7,5,4,3,5,3,4,5,4,3,
        5,7,4,6,6,6,5,4,6,4,6,8,4,4,6,4,4,8,5,6,5,4,4,4,
        5,6,4,8,4,8,5,4,5,6,6,6,4,6,8,6,5,4,6,4,5,6,5,4,
        6,8,5,7,7,7,6,5,7,5,7,9,5,5,7,5,5,9,6,7,6,5,5,5,
        6,7,5,9,5,9,6,5,6,7,7,7,5,7,9,7,6,5,7,5,6,7,6,5,
        7,9,6,8,8,8,7,6,8,6,8,10,6,6,8,6,6,10,7,8,7,6,6,6,
        7,8,6,10,6,10,7,6,7,8,8,8,6,8,10,8,7,6,8,6,7,8,7,6,
        8,10,7,9,9,9,8,7,9,7,9,11,7,7,9,7,7,11,8,9,8,7,7,7,
        8,9,7,11,7,11,8,7,8,9,9,9,7,9,11,9,8,7,9,7,8,9,8,7,
        9,11,8,10,10,10,9,8,10,8,10,12,8,8,10,8,8,12,9,10,9,8,8,8,
        9,10,8,12,8,12,9,8,9,10,10,10,8,10,12,10,9,8,10,8,9,10,9,8
    ], dtype=float),
    "Original McKenna": np.array([
        0,0,0,2,7,4,3,2,6,8,13,5,26,25,24,15,13,16,14,19,17,24,20,25,
        63,60,56,55,47,53,36,38,39,43,39,35,22,24,22,21,29,30,27,26,26,21,23,19,
        57,62,61,55,57,57,35,50,40,29,28,26,50,51,52,61,60,60,42,42,43,43,42,41,
        45,41,46,23,35,34,21,21,19,51,40,49,29,29,31,40,36,33,29,26,30,16,18,14,
        66,64,64,56,53,57,49,51,47,44,46,47,56,51,53,25,37,30,31,28,30,36,35,22,
        28,32,27,32,34,35,52,49,48,51,51,53,40,43,42,26,30,28,55,41,53,52,51,47,
        61,64,65,39,41,41,22,21,23,43,41,38,24,22,24,14,17,19,52,50,47,42,40,42,
        26,27,27,34,38,33,44,44,42,41,40,37,33,31,26,44,34,38,46,44,44,36,37,34,
        36,36,36,38,43,38,27,26,30,32,37,29,50,49,48,29,37,36,10,19,17,24,20,25,
        53,52,50,53,57,55,34,44,45,13,9,5,34,26,32,31,41,42,31,32,30,21,19,23,
        43,36,31,47,45,43,47,62,52,41,36,38,46,47,40,43,42,42,36,38,43,53,52,53,
        47,49,48,47,41,44,15,11,19,51,40,49,23,23,25,34,30,27,7,4,4,32,22,32,
        68,70,66,68,79,71,43,45,41,38,40,41,24,25,23,35,33,38,43,50,48,18,17,26,
        34,38,33,38,40,41,34,31,30,33,33,35,28,23,22,26,30,26,75,77,71,62,63,63,
        37,40,41,49,47,51,32,37,33,49,47,44,32,38,28,38,39,37,22,20,17,44,50,40,
        32,33,33,40,44,39,32,32,40,39,34,41,33,33,32,32,38,36,22,20,20,12,13,10
    ], dtype=float)
}

HEXAGRAM_NAMES = [
    "1. The Creative", "2. The Receptive", "3. Difficulty at the Beginning", "4. Youthful Folly",
    "5. Waiting", "6. Conflict", "7. The Army", "8. Holding Together",
    "9. The Taming Power of the Small", "10. Treading", "11. Peace", "12. Standstill",
    "13. Fellowship with Men", "14. Possession in Great Measure", "15. Modesty", "16. Enthusiasm",
    "17. Following", "18. Work on What Has Been Spoiled", "19. Approach", "20. Contemplation",
    "21. Biting Through", "22. Grace", "23. Splitting Apart", "24. Return",
    "25. Innocence", "26. The Taming Power of the Great", "27. The Corners of the Mouth", "28. Preponderance of the Great",
    "29. The Abysmal", "30. The Clinging", "31. Influence", "32. Duration",
    "33. Retreat", "34. The Power of the Great", "35. Progress", "36. Darkening of the Light",
    "37. The Family", "38. Opposition", "39. Obstruction", "40. Deliverance",
    "41. Decrease", "42. Increase", "43. Breakthrough", "44. Coming to Meet",
    "45. Gathering Together", "46. Pushing Upward", "47. Oppression", "48. The Well",
    "49. Revolution", "50. The Cauldron", "51. The Arousing", "52. Keeping Still",
    "53. Development", "54. The Marrying Maiden", "55. Abundance", "56. The Wanderer",
    "57. The Gentle", "58. The Joyous", "59. Dispersion", "60. Limitation",
    "61. Inner Truth", "62. Preponderance of the Small", "63. After Completion", "64. Before Completion"
]

HEXAGRAM_PATTERNS = [
    [1,1,1,1,1,1],  # 1:  Qian
    [0,0,0,0,0,0],  # 2:  Kun
    [1,0,0,0,1,0],  # 3:  Zhun
    [0,1,0,0,0,1],  # 4:  Meng
    [1,1,1,0,1,0],  # 5:  Xu
    [0,1,0,1,1,1],  # 6:  Song
    [0,0,0,0,1,0],  # 7:  Shi
    [0,1,0,0,0,0],  # 8:  Bi
    [1,1,1,0,1,1],  # 9:  Xiaoxu
    [1,1,0,1,1,1],  # 10: Lü
    [0,0,0,1,1,1],  # 11: Tai
    [1,1,1,0,0,0],  # 12: Pi
    [1,0,1,1,1,1],  # 13: Tongren
    [1,1,1,1,0,1],  # 14: Dayou
    [0,0,0,1,0,0],  # 15: Qian
    [0,0,1,0,0,0],  # 16: Yu
    [1,0,0,1,1,1],  # 17: Sui
    [1,1,1,0,0,1],  # 18: Gu
    [0,0,0,0,1,1],  # 19: Lin
    [1,1,0,0,0,0],  # 20: Guan
    [1,0,1,0,0,1],  # 21: Shihe
    [1,0,0,1,0,1],  # 22: Bi
    [1,0,0,0,0,0],  # 23: Bo
    [0,0,0,0,0,1],  # 24: Fu
    [1,0,0,1,1,1],  # 25: Wuwang
    [1,1,1,0,0,1],  # 26: Daxu
    [1,0,0,0,0,1],  # 27: Yi
    [0,1,1,1,1,0],  # 28: Daguo
    [0,1,0,0,1,0],  # 29: Kan
    [1,0,1,1,0,1],  # 30: Li
    [0,0,1,1,1,0],  # 31: Xian
    [0,1,1,1,0,0],  # 32: Heng
    [1,1,1,1,0,0],  # 33: Dun
    [0,0,1,1,1,1],  # 34: Dazhuang
    [0,1,0,1,0,0],  # 35: Jin
    [0,0,1,0,1,0],  # 36: Mingyi
    [1,0,1,0,1,1],  # 37: Jiaren
    [1,1,0,1,0,1],  # 38: Kui
    [0,1,0,1,0,0],  # 39: Jian
    [0,0,1,0,1,0],  # 40: Jie
    [1,1,0,0,0,1],  # 41: Sun
    [1,0,0,0,1,1],  # 42: Yi
    [1,1,1,1,1,0],  # 43: Guai
    [0,1,1,1,1,1],  # 44: Gou
    [0,0,0,1,1,0],  # 45: Cui
    [0,1,1,0,0,0],  # 46: Sheng
    [0,1,0,1,1,0],  # 47: Kun
    [0,1,1,0,1,0],  # 48: Jing
    [1,0,1,1,1,0],  # 49: Ge
    [0,1,1,1,0,1],  # 50: Ding
    [1,0,0,1,0,0],  # 51: Zhen
    [0,0,1,0,0,1],  # 52: Gen
    [0,0,1,0,1,1],  # 53: Jian
    [1,1,0,1,0,0],  # 54: Guimei
    [1,0,1,1,0,0],  # 55: Feng
    [0,0,1,1,0,1],  # 56: Lü
    [0,1,1,0,1,1],  # 57: Xun
    [1,1,0,1,1,0],  # 58: Dui
    [0,1,0,0,1,1],  # 59: Huan
    [1,1,0,0,1,0],  # 60: Jie
    [1,1,0,0,1,1],  # 61: Zhongfu
    [0,0,1,1,0,0],  # 62: Xiaoguo
    [1,0,1,0,1,0],  # 63: Jiji
    [0,1,0,1,0,1],  # 64: Weiji
]

# Built-in historical events: (label, CE year as float; negative = BCE)
HISTORICAL_EVENTS = [
    # ── Modern era ──────────────────────────────────────────────
    ("Fall of Rome",          476.0),
    ("Black Death",           1347.0),
    ("Gutenberg Press",       1450.0),
    ("Columbus",              1492.0),
    ("Newton Principia",      1687.0),
    ("French Revolution",     1789.0),
    ("Darwin Origin",         1859.0),
    ("WWI Start",             1914.0),
    ("WWII End",              1945.0),
    ("Hiroshima",             1945.61),
    ("Moon Landing",          1969.58),
    ("Internet / WWW",        1991.0),
    ("9/11",                  2001.69),
    ("Human Genome",          2003.0),
    ("2012 Solstice",         2012.97),
    ("COVID-19",              2020.17),
    # ── Classical antiquity ──────────────────────────────────────
    ("Jesus Birth (est.)",     0.0),
    ("Julius Caesar †",       -44.0),
    ("Fall of Carthage",      -146.0),
    ("Alexander the Great †", -323.0),
    ("Battle of Marathon",    -490.0),
    ("Socrates †",            -399.0),
    ("Confucius born",        -551.0),
    ("Birth of Buddha",       -563.0),
    ("Fall of Assyria",       -612.0),
    ("Homer (est.)",          -800.0),
    ("First Olympic Games",   -776.0),
    # ── Ancient world ────────────────────────────────────────────
    ("Iron Age begins",       -1200.0),
    ("Trojan War (est.)",     -1200.0),
    ("Exodus / Moses (est.)", -1250.0),
    ("Tutankhamun †",         -1323.0),
    ("Battle of Megiddo",     -1457.0),
    ("Code of Hammurabi",     -1754.0),
    ("Fall of Ur III",        -2004.0),
    ("Great Pyramid",         -2560.0),
    ("Stonehenge (phase 1)",  -3000.0),
    ("Sumerian writing",      -3200.0),
    ("Mayan Cal. Start",      -3114.0),
    ("Göbekli Tepe",          -9600.0),
    # ── Library of Alexandria ─────────────────────────────────────
    ("Alexandria founded",    -331.0),
    ("Library of Alexandria †", -48.0),
]

# Defaults used by "Reset to Defaults" button
DEFAULTS = {
    'primary_date':    '2012-12-21',
    'primary_era':     'CE',
    'secondary_date':  '',
    'secondary_era':   'CE',
    'hex_num':         1,
    'processed_hex':   1,
    'levels':          6,
    'years_before':    20,
    'months_before':   0,
    'days_before':     0,
    'years_after':     5,
    'months_after':    0,
    'days_after':      0,
    'invert':          True,
    'resolution':      1000,
    'threshold':       0.05,
    'marker_order':    5,
    'min_dist_days':   0,
    'show_resonance':  True,
    'selected_events': [],
    'custom_events':   '',
    'alignment_score':    None,
    'last_throw_hex':    None,
    'anim_resolution':   600,
    'anim_prim':         'Kelley',
    'anim_show_resonance': False,
    'anim_data':         None,
    # Navigation — persists active tab across reruns
    'active_tab':        'Kelley',
    'math_active_tab':   'Equations',
    'manual_active_tab': 'Theory',
    # Math tab widgets — key= only, no value=
    'calc_prim':         'Kelley',
    'calc_hex':          1,
    'calc_levels':       6,
    'calc_t':            0.0,
    'xl_prim':           'Kelley',
    'xl_hex':            1,
    'xl_levels':         5,
    'xl_points':         200,
    'xl_before':         7305,
    'xl_after':          1826,
}

# =============================================================================
# CORE MATH
# =============================================================================

def novelty_function(x_days, levels, primitive):
    """Fractal novelty accumulator with clean overflow handling."""
    y = np.zeros_like(x_days, dtype=float)
    scale = 1.0
    with np.errstate(over='ignore', invalid='ignore'):
        for _ in range(levels):
            frac = np.mod(x_days / scale, 1.0)
            idx  = np.floor(x_days / scale).astype(int) % 384
            v1   = primitive[idx]
            v2   = primitive[(idx + 1) % 384]
            v    = v1 + frac * (v2 - v1)
            y   += v * scale
            scale *= 64.0
    # Clamp infinities to the finite range found in the array
    finite_mask = np.isfinite(y)
    if finite_mask.any():
        ymax = y[finite_mask].max()
        ymin = y[finite_mask].min()
    else:
        ymax, ymin = 1e15, -1e15
    y = np.nan_to_num(y, nan=0.0, posinf=ymax, neginf=ymin)
    return y

def calculate_timewave(primitive_name, levels, hex_num, x_days, invert=True):
    primitive = PRIMITIVES[primitive_name].copy()
    primitive = np.roll(primitive, -(hex_num - 1) * 6)
    y = novelty_function(x_days, levels, primitive)
    return -y if invert else y

@st.cache_data(show_spinner=False)
def compute_all_waves(levels, hex_num, days_before_total, days_after_total, invert, resolution):
    """Heavy computation — cached by all input parameters."""
    x_days = np.linspace(-days_before_total, days_after_total, resolution)
    result = {}
    for prim in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
        result[prim] = calculate_timewave(prim, levels, hex_num, x_days, invert)
    return x_days, result

@st.cache_data(show_spinner=False)
def compute_resonance_markers(y_tuple, dates_tuple, threshold, order, min_dist_days, x_days_tuple):
    """Cached resonance detection with order and minimum-distance filtering."""
    y_values  = np.array(y_tuple)
    dates_str = list(dates_tuple)
    x_days    = np.array(x_days_tuple)

    peaks_list, valleys_list = [], []
    if len(y_values) < max(3, order * 2 + 1):
        return peaks_list, valleys_list

    dy = np.diff(y_values)
    dy_norm = dy / (np.max(np.abs(dy)) + 1e-9)

    peaks_idx   = argrelextrema(y_values, np.greater, order=order)[0]
    valleys_idx = argrelextrema(y_values, np.less,    order=order)[0]

    def _filter(indices, is_peak):
        accepted = []
        last_x   = -np.inf
        for idx in indices:
            if not (0 < idx < len(y_values) - 1):
                continue
            sharp = abs(dy_norm[idx-1]) > threshold or abs(dy_norm[idx]) > threshold
            if not sharp:
                continue
            x_val = float(x_days[idx])
            if min_dist_days > 0 and (x_val - last_x) < min_dist_days:
                continue
            accepted.append((dates_str[idx], float(y_values[idx])))
            last_x = x_val
        return accepted

    peaks_list   = _filter(peaks_idx,   is_peak=True)
    valleys_list = _filter(valleys_idx, is_peak=False)
    return peaks_list, valleys_list

# compute_animation_frames removed in v3.2:
# animation is now built inline in the tab with st.progress feedback.
# The underlying calculate_timewave is cached, so per-level calls are fast.

# =============================================================================
# DATE HANDLING
# =============================================================================

_EPOCH = pd.Timestamp("1970-01-01")

def _days_in_month(year: int, month: int) -> int:
    """Return number of days in month, leap-year-aware."""
    if year < 1:
        # Approximate for BCE — use proleptic Gregorian rule
        year_for_check = abs(year)
        is_leap = (year_for_check % 4 == 0 and
                   (year_for_check % 100 != 0 or year_for_check % 400 == 0))
        if month == 2:
            return 29 if is_leap else 28
        return [31,28,31,30,31,30,31,31,30,31,30,31][month-1]
    return calendar.monthrange(year, month)[1]

def parse_date_string(date_str: str):
    """
    Parse CE and BCE date strings.
    Returns pd.Timestamp for CE, dict for BCE.
    Raises ValueError with user-friendly message on bad input.
    """
    date_str = date_str.strip()
    if not date_str:
        raise ValueError("Date is empty. Use YYYY-MM-DD (e.g. 2012-12-21).")

    if date_str.upper().startswith('BCE'):
        date_str = '-' + date_str[3:].strip()

    if date_str.startswith('-'):
        parts = date_str[1:].split('-')
        if len(parts) != 3:
            raise ValueError("BCE format: BCE YYYY-MM-DD  (e.g. BCE 3114-08-11)")
        try:
            year  = -int(parts[0])
            month = int(parts[1])
            day   = int(parts[2])
        except ValueError:
            raise ValueError("BCE date parts must be integers: BCE YYYY-MM-DD")
        if not (1 <= month <= 12):
            raise ValueError(f"Month must be 1–12, got {month}")
        max_day = _days_in_month(year, month)
        if not (1 <= day <= max_day):
            raise ValueError(f"{abs(year)} BCE, month {month}: day must be 1–{max_day}, got {day}")
        return {'year': year, 'month': month, 'day': day, 'is_bce': True}
    else:
        # CE date — let pandas handle validation (it raises on Feb 30 etc.)
        try:
            ts = pd.Timestamp(date_str)
            if pd.isna(ts):
                raise ValueError("Empty or unparseable date.")
        except Exception:
            raise ValueError(
                f"Cannot parse '{date_str}'. Use YYYY-MM-DD (e.g. 2012-12-21). "
                "Check for invalid days like Feb 29 in a non-leap year."
            )
        return ts

def date_to_days_from_epoch(date_obj) -> int:
    """Days from Unix epoch (1970-01-01). BCE uses 365.25 approximation."""
    if isinstance(date_obj, dict) and date_obj.get('is_bce'):
        year  = date_obj['year']   # negative
        days  = (year - 1970) * 365.25 + (date_obj['month'] - 1) * 30.4375 + date_obj['day']
        return int(days)
    elif isinstance(date_obj, pd.Timestamp):
        return (date_obj - _EPOCH).days
    elif isinstance(date_obj, datetime):
        return (pd.Timestamp(date_obj) - _EPOCH).days
    return 0

def _epoch_days_to_proleptic_gregorian(total_days: int):
    """
    Convert integer days from Unix epoch (1970-01-01) to a proleptic Gregorian
    (year, month, day). Works for ANY year with no overflow.
    Uses Howard Hinnant's civil calendar algorithm (public domain).
    Astronomical year numbering: year 0 = 1 BCE, year -1 = 2 BCE.
    """
    z   = int(total_days) + 719468
    era = (z if z >= 0 else z - 146096) // 146097
    doe = z - era * 146097
    yoe = (doe - doe // 1460 + doe // 36524 - doe // 146096) // 365
    y   = yoe + era * 400
    doy = doe - (365 * yoe + yoe // 4 - yoe // 100)
    mp  = (5 * doy + 2) // 153
    d   = doy - (153 * mp + 2) // 5 + 1
    m   = mp + (3 if mp < 10 else -9)
    y  += (1 if m <= 2 else 0)
    return int(y), int(m), int(d)

def _format_proleptic(year: int, month: int, day: int) -> str:
    """Format proleptic Gregorian date. year<=0 → BCE label, year>0 → CE label."""
    if year <= 0:
        bce_year = 1 - year
        return f"BCE {bce_year:04d}-{month:02d}-{day:02d}"
    return f"{year:04d}-{month:02d}-{day:02d}"

def generate_date_labels(zero_date, x_days):
    """
    Convert x_days offsets (days from zero_date) to display strings.
    Uses pure-arithmetic proleptic Gregorian — no pd.Timestamp overflow.
    Works correctly for any range: BCE 9999 to CE 9999+.
    Both CE and BCE zero dates use the same code path.
    """
    zero_epoch = date_to_days_from_epoch(zero_date)
    dates = []
    for offset in x_days:
        epoch_day = zero_epoch + int(round(float(offset)))
        y, m, d = _epoch_days_to_proleptic_gregorian(epoch_day)
        dates.append(_format_proleptic(y, m, d))
    return dates

def calculate_date_range_offset(years, months, days) -> int:
    return int(years * 365.25 + months * 30.4375 + days)

def format_zero_date_for_display(zero_date_obj) -> str:
    if isinstance(zero_date_obj, dict) and zero_date_obj.get('is_bce'):
        return f"BCE {abs(zero_date_obj['year']):04d}-{zero_date_obj['month']:02d}-{zero_date_obj['day']:02d}"
    elif isinstance(zero_date_obj, pd.Timestamp):
        return zero_date_obj.strftime('%Y-%m-%d')
    return str(zero_date_obj)

def event_year_to_x_offset(event_year_float: float, zero_date) -> float:
    """Return x_days offset (days from zero_date) for a CE float year."""
    zero_days  = date_to_days_from_epoch(zero_date)
    event_days = (event_year_float - 1970) * 365.25
    return event_days - zero_days

def parse_custom_events(text: str) -> list:
    """
    Parse user-supplied custom events from a text area.
    Each line: Label | YYYY-MM-DD   OR   Label | BCE YYYY-MM-DD
    Lines with errors are silently skipped and reported back.
    Returns list of (label, year_float) tuples and a list of error strings.
    """
    events = []
    errors = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        if '|' not in line:
            errors.append(f"Skipped (no '|'): {line!r}")
            continue
        parts = line.split('|', 1)
        label = parts[0].strip()
        date_part = parts[1].strip()
        try:
            parsed = parse_date_string(date_part)
            if isinstance(parsed, dict):
                year_float = float(parsed['year']) + (parsed['month'] - 1) / 12.0
            else:
                year_float = parsed.year + (parsed.month - 1) / 12.0 + (parsed.day - 1) / 365.25
            events.append((label, year_float))
        except Exception as e:
            errors.append(f"Skipped '{label}': {e}")
    return events, errors

# =============================================================================
# ALIGNMENT SCORE
# =============================================================================

def compute_alignment_score(resonance_data, all_events, zero_date, x_days, dates_str, window_days=365):
    """
    For each event in range, check if any resonance marker (across all primitives)
    falls within window_days of it.  Returns (score, possible, matched_list).
    """
    # Build resonance x-offsets set from dates
    date_to_idx = {d: i for i, d in enumerate(dates_str)}
    res_x_set = []
    for prim_data in resonance_data.values():
        for date_s, _ in prim_data.get('peaks', []) + prim_data.get('valleys', []):
            if date_s in date_to_idx:
                res_x_set.append(x_days[date_to_idx[date_s]])
    res_x_arr = np.array(res_x_set) if res_x_set else np.array([])

    event_offsets = []
    for label, year_float in all_events:
        offset = event_year_to_x_offset(year_float, zero_date)
        if x_days[0] <= offset <= x_days[-1]:
            event_offsets.append((label, offset))

    if not event_offsets or len(res_x_arr) == 0:
        return 0, len(event_offsets), []

    matched = []
    for label, off in event_offsets:
        dists = np.abs(res_x_arr - off)
        if dists.min() <= window_days:
            matched.append((label, float(dists.min())))

    score = len(matched)
    return score, len(event_offsets), matched

# =============================================================================
# PLOTTING HELPERS
# =============================================================================

def draw_hexagram(hexagram_num):
    """Return PNG bytes for a hexagram glyph. Using st.image(bytes) avoids
    the Streamlit media-file-storage cache expiry bug that occurs with st.pyplot()."""
    import io as _io
    pattern = HEXAGRAM_PATTERNS[hexagram_num - 1]
    fig, ax = plt.subplots(figsize=(0.8, 1.4))
    ax.set_xlim(0, 1)
    ax.set_ylim(-0.5, 5.5)
    ax.axis('off')
    for line_num in range(6):
        y_pos = line_num
        bit   = pattern[line_num]
        if bit == 1:
            ax.plot([0.15, 0.85], [y_pos, y_pos], color='black', linewidth=7, solid_capstyle='butt')
        else:
            ax.plot([0.15, 0.42], [y_pos, y_pos], color='black', linewidth=7, solid_capstyle='butt')
            ax.plot([0.58, 0.85], [y_pos, y_pos], color='black', linewidth=7, solid_capstyle='butt')
    fig.tight_layout(pad=0.1)
    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=90, bbox_inches='tight', transparent=True)
    plt.close(fig)
    buf.seek(0)
    return buf.read()

def add_zero_vlines(fig, zero_date, zero_date2, x_days):
    zero_x = float(x_days[int(np.argmin(np.abs(x_days)))])
    fig.add_vline(x=zero_x, line_dash="dash", line_color="red", line_width=2,
                  annotation_text="Primary Zero", annotation_position="top right")
    if zero_date2 is not None:
        offset = float(date_to_days_from_epoch(zero_date2) - date_to_days_from_epoch(zero_date))
        if x_days[0] <= offset <= x_days[-1]:
            fig.add_vline(x=offset, line_dash="dot", line_color="purple", line_width=2,
                          annotation_text="Secondary Zero", annotation_position="bottom right")

def add_event_overlays(fig, all_events, zero_date, x_days, y_values=None):
    """
    Overlay built-in + custom events as vertical lines with hoverable tooltips.
    Each line gets an invisible Scatter point at its base so Plotly shows a
    tooltip with the event label and exact calendar date on hover.
    """
    zero_epoch = date_to_days_from_epoch(zero_date)
    # Place hover points at the bottom of the data range
    y_base = float(np.nanmin(y_values)) if y_values is not None and len(y_values) > 0 else 0.0

    for label, year_float in all_events:
        offset = event_year_to_x_offset(year_float, zero_date)
        if x_days[0] <= offset <= x_days[-1]:
            x_val = float(offset)
            # Compute exact calendar date for this offset
            epoch_day = zero_epoch + int(round(x_val))
            yr, mo, dy = _epoch_days_to_proleptic_gregorian(epoch_day)
            exact_date = _format_proleptic(yr, mo, dy)
            # Visible vertical line
            fig.add_vline(x=x_val,
                          line_dash="longdash", line_color="rgba(80,80,80,0.50)",
                          line_width=1,
                          annotation_text=label, annotation_position="top left",
                          annotation_font_size=9)
            # Invisible Scatter point for tooltip
            fig.add_trace(go.Scatter(
                x=[x_val], y=[y_base],
                mode='markers',
                marker=dict(size=14, opacity=0, color='rgba(0,0,0,0)'),
                name=label, showlegend=False,
                hovertemplate=(
                    f"<b>{label}</b><br>"
                    f"Date: {exact_date}<extra></extra>"
                )
            ))

def add_resonance_traces(fig, peaks, valleys, x_days, dates_str):
    date_to_x = {d: x_days[i] for i, d in enumerate(dates_str)}
    pk_x  = [date_to_x[d] for d, _ in peaks   if d in date_to_x]
    pk_y  = [v             for d, v in peaks   if d in date_to_x]
    vl_x  = [date_to_x[d] for d, _ in valleys if d in date_to_x]
    vl_y  = [v             for d, v in valleys if d in date_to_x]
    pk_cd = [d             for d, _ in peaks   if d in date_to_x]
    vl_cd = [d             for d, _ in valleys if d in date_to_x]
    if pk_x:
        fig.add_trace(go.Scatter(
            x=pk_x, y=pk_y, mode='markers',
            marker=dict(color='limegreen', size=8, symbol='triangle-up',
                        line=dict(color='darkgreen', width=1)),
            name='Peaks', customdata=pk_cd,
            hovertemplate='Peak<br>%{customdata}<br>%{y:.4f}<extra></extra>'
        ))
    if vl_x:
        fig.add_trace(go.Scatter(
            x=vl_x, y=vl_y, mode='markers',
            marker=dict(color='tomato', size=8, symbol='triangle-down',
                        line=dict(color='darkred', width=1)),
            name='Valleys', customdata=vl_cd,
            hovertemplate='Valley<br>%{customdata}<br>%{y:.4f}<extra></extra>'
        ))

def make_x_axis_layout(x_days, dates_str, n_ticks=12):
    step     = max(1, len(x_days) // n_ticks)
    tick_v   = x_days[::step].tolist()
    tick_t   = [dates_str[i] for i in range(0, len(dates_str), step)]
    return dict(tickvals=tick_v, ticktext=tick_t, tickangle=35, title="Date")

# =============================================================================
# URL STATE
# =============================================================================

def load_url_params() -> dict:
    p = st.query_params
    def _b(key, default):
        v = p.get(key, str(default))
        return v.lower() in ('1', 'true', 'yes')
    return {
        'primary_date':    p.get('pd',  DEFAULTS['primary_date']),
        'primary_era':     p.get('pe',  DEFAULTS['primary_era']),
        'secondary_date':  p.get('sd',  DEFAULTS['secondary_date']),
        'secondary_era':   p.get('se',  DEFAULTS['secondary_era']),
        'hex_num':         int(p.get('hx',  DEFAULTS['hex_num'])),
        'levels':          int(p.get('lv',  DEFAULTS['levels'])),
        'years_before':    int(p.get('yb',  DEFAULTS['years_before'])),
        'months_before':   int(p.get('mb',  DEFAULTS['months_before'])),
        'days_before':     int(p.get('db',  DEFAULTS['days_before'])),
        'years_after':     int(p.get('ya',  DEFAULTS['years_after'])),
        'months_after':    int(p.get('ma',  DEFAULTS['months_after'])),
        'days_after':      int(p.get('da',  DEFAULTS['days_after'])),
        'invert':          _b('inv', DEFAULTS['invert']),
        'resolution':      int(p.get('res', DEFAULTS['resolution'])),
        'threshold':       float(p.get('thr', DEFAULTS['threshold'])),
        'show_resonance':  _b('sr',  DEFAULTS['show_resonance']),
    }

def save_url_params(**kwargs):
    mapping = {
        'primary_date':   'pd',  'primary_era':  'pe',
        'secondary_date': 'sd',  'secondary_era':'se',
        'hex_num':        'hx',  'levels':       'lv',
        'years_before':   'yb',  'months_before':'mb', 'days_before': 'db',
        'years_after':    'ya',  'months_after': 'ma', 'days_after':  'da',
        'invert':         'inv', 'resolution':   'res','threshold':   'thr',
        'show_resonance': 'sr',
    }
    updates = {}
    for long, short in mapping.items():
        if long in kwargs:
            v = kwargs[long]
            updates[short] = '1' if v is True else ('0' if v is False else str(v))
    st.query_params.update(updates)

# =============================================================================
# SESSION STATE INIT
# =============================================================================

url_d = load_url_params()

def _ss(key, fallback=None):
    default = url_d.get(key, DEFAULTS.get(key, fallback))
    if key not in st.session_state:
        st.session_state[key] = default

for k in DEFAULTS:
    _ss(k)
# processed_hex mirrors hex_num on first load
if 'processed_hex' not in st.session_state:
    st.session_state.processed_hex = st.session_state.hex_num
if 'hex_just_processed' not in st.session_state:
    st.session_state.hex_just_processed = False

st.markdown(
    '<h1 style="display:flex;align-items:center;gap:12px;">'
    '<span style="background:#1565C0;color:white;border-radius:8px;'
    'padding:4px 10px;font-size:1.1em;">☯</span>'
    ' Novelty Explorer'
    '</h1>',
    unsafe_allow_html=True
)

# =============================================================================
# ONBOARDING — moved to Manual tab (see below)
# =============================================================================

# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.header("⚙️ Configuration")

    # ── Levels & Hexagram ──────────────────────────────────────────────────
    with st.expander("🎲 Levels & Hexagram", expanded=True):
        levels = st.number_input("Levels (1–16)", 1, 16,
                                 value=st.session_state.levels, step=1, key='levels_input')
        st.session_state.levels = levels

        if levels > 13:
            st.warning(f"⚠️ Level {levels}: scale ≈ 64^{levels} — overflow clamped automatically.")

        hex_num = st.number_input("Hexagram (1–64)", 1, 64,
                                  value=st.session_state.hex_num, step=1, key='hex_input')
        st.markdown(
            '<p style="color:#E07B00;font-weight:bold;font-size:0.9em;">'
            '▲ Press Process to apply the selected hexagram'
            '</p>',
            unsafe_allow_html=True
        )

        col_throw, col_proc = st.columns(2)
        with col_throw:
            if st.button("🎲 Throw", use_container_width=True):
                new_hex = random.randint(1, 64)
                st.session_state.hex_num         = new_hex
                st.session_state.processed_hex   = new_hex
                st.session_state.last_throw_hex  = new_hex
                st.session_state.alignment_score = None
                st.session_state.hex_just_processed = True  # show green banner
        with col_proc:
            if st.button("✓ Process", use_container_width=True):
                st.session_state.hex_num         = hex_num
                st.session_state.processed_hex   = hex_num
                st.session_state.hex_just_processed = True  # show green banner

        # Green confirmation banner — persists until the next Process or Throw
        if st.session_state.hex_just_processed:
            hex_name = HEXAGRAM_NAMES[st.session_state.processed_hex - 1]
            st.success(
                f"✓ Hexagram **{st.session_state.processed_hex} — {hex_name}** "
                f"is active"
            )
        else:
            st.caption(f"Active: **Hex {st.session_state.processed_hex}** — "
                       f"{HEXAGRAM_NAMES[st.session_state.processed_hex - 1]}")

    # ── Primary Zero Date ─────────────────────────────────────────────────
    with st.expander("📅 Primary Zero Date", expanded=True):
        primary_date_input = st.text_input("Date (YYYY-MM-DD)",
                                           value=st.session_state.primary_date,
                                           key="primary_date_input")
        st.session_state.primary_date = primary_date_input

        primary_era = st.radio("Era", ["CE", "BCE"],
                               index=0 if st.session_state.primary_era == "CE" else 1,
                               key="primary_era_radio", horizontal=True)
        st.session_state.primary_era = primary_era

        st.caption("Quick presets:")
        pc1, pc2 = st.columns(2)
        with pc1:
            if st.button("2012-12-21", use_container_width=True, key="preset_2012"):
                st.session_state.primary_date = "2012-12-21"
                st.session_state.primary_era  = "CE"
        with pc2:
            if st.button("BCE 3114-08-11", use_container_width=True, key="preset_mayan"):
                st.session_state.primary_date = "3114-08-11"
                st.session_state.primary_era  = "BCE"

    # ── Secondary Zero Date ───────────────────────────────────────────────
    with st.expander("📅 Secondary Zero Date (optional)", expanded=False):
        secondary_date_input = st.text_input("Date (YYYY-MM-DD)",
                                             value=st.session_state.secondary_date,
                                             key="secondary_date_input",
                                             placeholder="Leave blank to disable")
        st.session_state.secondary_date = secondary_date_input

        secondary_era = st.radio("Era", ["CE", "BCE"],
                                 index=0 if st.session_state.secondary_era == "CE" else 1,
                                 key="secondary_era_radio", horizontal=True)
        st.session_state.secondary_era = secondary_era

    # ── Date Range ────────────────────────────────────────────────────────
    with st.expander("📏 Date Range", expanded=True):
        st.markdown("**Before Zero Date**")
        rc1, rc2, rc3 = st.columns(3)
        with rc1:
            years_before  = st.number_input("Yrs",  0, 4000, value=st.session_state.years_before,  key='yb')
            st.session_state.years_before = years_before
        with rc2:
            months_before = st.number_input("Mo",   0, 11,   value=st.session_state.months_before, key='mb')
            st.session_state.months_before = months_before
        with rc3:
            days_before   = st.number_input("Days", 0, 365,  value=st.session_state.days_before,   key='db')
            st.session_state.days_before = days_before

        st.markdown("**After Zero Date**")
        ra1, ra2, ra3 = st.columns(3)
        with ra1:
            years_after   = st.number_input("Yrs",  0, 3000, value=st.session_state.years_after,  key='ya')
            st.session_state.years_after = years_after
        with ra2:
            months_after  = st.number_input("Mo",   0, 11,   value=st.session_state.months_after, key='ma')
            st.session_state.months_after = months_after
        with ra3:
            days_after    = st.number_input("Days", 0, 365,  value=st.session_state.days_after,   key='da')
            st.session_state.days_after = days_after

    # ── Display Options ───────────────────────────────────────────────────
    with st.expander("🖥️ Display Options", expanded=True):
        invert = st.checkbox("Invert Novelty", value=st.session_state.invert)
        st.session_state.invert = invert

        resolution = st.slider("Plot Resolution", 400, 2400,
                                value=st.session_state.resolution, step=100,
                                help="Points per plot. More = detail; fewer = speed.")
        st.session_state.resolution = resolution

        show_resonance = st.checkbox("Show Resonance Markers on plots",
                                     value=st.session_state.show_resonance,
                                     key="show_resonance_global")
        st.session_state.show_resonance = show_resonance

    # ── Resonance Tuning ─────────────────────────────────────────────────
    with st.expander("🎚️ Resonance Tuning", expanded=False):
        threshold = st.slider("Sharpness Threshold", 0.01, 0.30,
                              value=st.session_state.threshold, step=0.01,
                              help="Min normalised slope to qualify as a marker.")
        st.session_state.threshold = threshold

        marker_order = st.slider("Extrema Order (sensitivity)", 2, 30,
                                 value=st.session_state.marker_order, step=1,
                                 help="Higher = only prominent peaks/valleys. Lower = more sensitive.")
        st.session_state.marker_order = marker_order

        # Guard: argrelextrema needs ≥ 2*order+1 points either side.
        # Cap effective order to avoid silent zero-marker results.
        max_safe_order = max(2, st.session_state.resolution // 10)
        effective_order = min(marker_order, max_safe_order)
        if marker_order > max_safe_order:
            st.warning(
                f"⚠️ Order {marker_order} exceeds safe limit for resolution "
                f"{st.session_state.resolution} — capped to {effective_order}. "
                "Raise resolution or lower order to use full sensitivity."
            )

        min_dist_days = st.slider("Min Distance Between Markers (days)", 0, 3650,
                                  value=st.session_state.min_dist_days, step=10,
                                  help="Suppress markers closer together than this (0 = off).")
        st.session_state.min_dist_days = min_dist_days

    # ── Historical Event Overlays ─────────────────────────────────────────
    with st.expander("🏛️ Historical Event Overlays", expanded=False):
        all_builtin_labels = [label for label, _ in HISTORICAL_EVENTS]

        selected_builtin = st.multiselect("Built-in events", all_builtin_labels,
                                          default=st.session_state.selected_events,
                                          key='selected_events_widget')
        st.session_state.selected_events = selected_builtin

        st.markdown("**Custom Events**")
        st.caption(
            "One per line — format: `Label | YYYY-MM-DD` or `Label | BCE YYYY-MM-DD`  \n"
            "Examples: `Moon Landing | 1969-07-20`   `Great Pyramid | BCE 2560-01-01`"
        )
        custom_text = st.text_area("Custom event list", value=st.session_state.custom_events,
                                   height=130, key='custom_events_widget',
                                   placeholder="Moon Landing | 1969-07-20\nGreat Pyramid | BCE 2560-01-01\nAI Singularity | 2045-01-01")
        st.session_state.custom_events = custom_text

    # ── Alignment Score ───────────────────────────────────────────────────
    with st.expander("🎯 Novelty Alignment Score", expanded=False):
        st.caption("After a 🎲 Throw, measure how often resonance markers land near historical events.")
        align_window = st.slider("Match window (days)", 30, 3650, 365, 30,
                                 help="A resonance within this many days of an event counts as a match.")
        if st.button("⚡ Compute Score", use_container_width=True):
            st.session_state.alignment_score = "PENDING"

    # ── Share & Reset ─────────────────────────────────────────────────────
    st.markdown("---")
    col_share, col_reset = st.columns(2)
    with col_share:
        if st.button("🔗 Share URL", use_container_width=True):
            save_url_params(
                primary_date   = st.session_state.primary_date,
                primary_era    = st.session_state.primary_era,
                secondary_date = st.session_state.secondary_date,
                secondary_era  = st.session_state.secondary_era,
                hex_num        = st.session_state.processed_hex,
                levels         = st.session_state.levels,
                years_before   = years_before,
                months_before  = months_before,
                days_before    = days_before,
                years_after    = years_after,
                months_after   = months_after,
                days_after     = days_after,
                invert         = invert,
                resolution     = resolution,
                threshold      = threshold,
                show_resonance = show_resonance,
            )
            st.success("URL updated — copy from address bar.")
    with col_reset:
        if st.button("↺ Reset", use_container_width=True):
            for k, v in DEFAULTS.items():
                st.session_state[k] = v
            st.rerun()

    st.caption("Timewave Zero · Terence McKenna & Peter Meyer · version 1_2026 · Streamlit + Plotly")

# =============================================================================
# BUILD ZERO DATE STRINGS FROM SESSION STATE
# =============================================================================

primary_era = st.session_state.primary_era
zero_date_str = (f"-{st.session_state.primary_date}"
                 if primary_era == "BCE"
                 else st.session_state.primary_date)

secondary_date_input = st.session_state.secondary_date
secondary_era        = st.session_state.secondary_era
if secondary_date_input.strip():
    zero_date2_str = (f"-{secondary_date_input}" if secondary_era == "BCE" else secondary_date_input)
else:
    zero_date2_str = ""

days_before_total = calculate_date_range_offset(years_before, months_before, days_before)
days_after_total  = calculate_date_range_offset(years_after,  months_after,  days_after)

# ── Zero-range guard ──────────────────────────────────────────────────────────
if days_before_total == 0 and days_after_total == 0:
    st.error("⛔ Set at least some range before or after the zero date (e.g. 1 year before).")
    st.stop()

# =============================================================================
# PARSE + COMPUTE
# =============================================================================

try:
    zero_date = parse_date_string(zero_date_str)
except Exception as e:
    st.error(f"❌ Primary date error: {e}")
    st.stop()

try:
    with st.spinner("Computing timewaves…"):
        x_days, all_y_values = compute_all_waves(
            st.session_state.levels,
            st.session_state.processed_hex,
            days_before_total, days_after_total,
            invert, resolution
        )
    dates_str = generate_date_labels(zero_date, x_days)
except Exception as e:
    st.error(f"❌ Computation error: {e}")
    st.stop()

zero_date2 = None
if zero_date2_str:
    try:
        zero_date2 = parse_date_string(zero_date2_str)
    except Exception as e:
        st.warning(f"⚠️ Secondary date ignored: {e}")

# ── Parse custom events ───────────────────────────────────────────────────────
custom_events_parsed, custom_errors = parse_custom_events(st.session_state.custom_events)
if custom_errors:
    with st.expander(f"⚠️ {len(custom_errors)} custom event(s) skipped — click to see", expanded=False):
        for err in custom_errors:
            st.caption(err)

# Combined events to display: built-in selected + custom
active_builtin  = [(l, y) for l, y in HISTORICAL_EVENTS if l in st.session_state.selected_events]
all_active_events = active_builtin + custom_events_parsed

# =============================================================================
# RESONANCE
# =============================================================================

resonance_data = {}
for prim_name in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
    y = all_y_values[prim_name]
    peaks, valleys = compute_resonance_markers(
        tuple(y.tolist()), tuple(dates_str),
        threshold, effective_order, min_dist_days,
        tuple(x_days.tolist())
    )
    resonance_data[prim_name] = {'peaks': peaks, 'valleys': valleys}

# ── Alignment Score computation ───────────────────────────────────────────────
all_events_for_score = list(HISTORICAL_EVENTS) + custom_events_parsed

if st.session_state.alignment_score == "PENDING":
    score, possible, matched = compute_alignment_score(
        resonance_data, all_events_for_score, zero_date, x_days, dates_str, align_window
    )
    st.session_state.alignment_score = (score, possible, matched)

if isinstance(st.session_state.alignment_score, tuple):
    score, possible, matched = st.session_state.alignment_score
    pct = (score / possible * 100) if possible else 0
    badge_col, _ = st.columns([1, 3])
    with badge_col:
        if pct >= 70:
            colour = "🟢"
        elif pct >= 40:
            colour = "🟡"
        else:
            colour = "🔴"
        st.metric(
            label=f"{colour} Alignment Score (Hex {st.session_state.processed_hex})",
            value=f"{score} / {possible}",
            delta=f"{pct:.0f}% of events matched (±{align_window}d window)"
        )
        if matched:
            with st.expander("Matched events"):
                for label, dist in sorted(matched, key=lambda x: x[1]):
                    st.caption(f"• {label} — nearest resonance {dist:.0f} days away")

# Pre-build shared x-axis layout
x_axis_layout = make_x_axis_layout(x_days, dates_str)

# =============================================================================
# TABS
# =============================================================================

# ── Persistent tab navigation via radio (st.tabs resets to tab-0 on every rerun) ──
_TAB_NAMES = ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna",
              "Comparison", "Resonance", "Animation", "Hexagrams", "Manual", "Math"]
st.markdown(
    '<style>'
    ' div[data-testid="stRadio"] > div { flex-wrap: wrap; gap: 4px; }'
    ' div[data-testid="stRadio"] label { '
    '   background:#AED6F1; color:#1a1a1a; border-radius:15px;'
    '   padding:4px 10px; cursor:pointer; font-size:0.85em; }'
    ' div[data-testid="stRadio"] label:has(input:checked) {'
    '   background:#D5D8DC; color:white; font-weight:bold; }'
    '</style>', unsafe_allow_html=True
)
active_tab = st.radio(
    "##nav", _TAB_NAMES, horizontal=True,
    key="active_tab", label_visibility="collapsed"
)

# ------------------------------------------------------------------
# Individual primitive tabs (0–3)
# ------------------------------------------------------------------
for prim_name in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
    if active_tab == prim_name:
        y       = all_y_values[prim_name]
        peaks   = resonance_data[prim_name]['peaks']
        valleys = resonance_data[prim_name]['valleys']

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=x_days, y=y,
            mode='lines', name=prim_name,
            line=dict(color=COLORS_PRIMITIVES[prim_name], width=2),
            hovertemplate='%{customdata}<br>Novelty: %{y:.4f}<extra></extra>',
            customdata=dates_str
        ))

        add_zero_vlines(fig, zero_date, zero_date2, x_days)
        add_event_overlays(fig, all_active_events, zero_date, x_days, y)
        if show_resonance:
            add_resonance_traces(fig, peaks, valleys, x_days, dates_str)

        fig.update_layout(
            title=(f"{prim_name} — Level {st.session_state.levels}, "
                   f"Hex {st.session_state.processed_hex}: "
                   f"{HEXAGRAM_NAMES[st.session_state.processed_hex - 1]}"),
            xaxis=x_axis_layout,
            yaxis_title="Novelty (inverted)" if invert else "Novelty",
            height=520, hovermode="x unified"
        )
        st.plotly_chart(fig, width='stretch')

        st.markdown(
            f"**Stats:** Avg `{np.mean(y):.2f}` | Min `{np.min(y):.2f}` | "
            f"Max `{np.max(y):.2f}` | Points `{len(y)}`"
        )
        csv_body = "Date,Days_Offset,Novelty\n" + "\n".join(
            f"{d},{x:.2f},{v:.6f}" for d, x, v in zip(dates_str, x_days, y)
        )
        st.download_button(f"📥 Download CSV", csv_body,
                           f"timewave_{prim_name}.csv", key=f"dl_{prim_name}")

# ------------------------------------------------------------------
# Comparison tab (5)
# ------------------------------------------------------------------
if active_tab == "Comparison":
    st.subheader("All Five Primitives — Comparison")
    st.caption("↓ scroll down for Correlation Matrix, Difference Overlay and data export")

    fig = go.Figure()
    for prim_name in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
        fig.add_trace(go.Scatter(
            x=x_days, y=all_y_values[prim_name],
            mode='lines', name=prim_name,
            line=dict(color=COLORS_PRIMITIVES[prim_name], width=1.5),
            hovertemplate='%{customdata}<br>' + prim_name + ': %{y:.4f}<extra></extra>',
            customdata=dates_str
        ))

    add_zero_vlines(fig, zero_date, zero_date2, x_days)
    add_event_overlays(fig, all_active_events, zero_date, x_days, all_y_values['Kelley'])
    if show_resonance:
        # Add resonance markers for ALL primitives, each in its own colour
        date_to_x = {d: x_days[i] for i, d in enumerate(dates_str)}
        _peak_colors   = {'Kelley':'limegreen','Watkins':'#F5A623','Sheliak':'#7ED321',
                          'HuangTi':'#E91E63','Original McKenna':'#CE93D8'}
        _valley_colors = {'Kelley':'tomato',   'Watkins':'#D0021B','Sheliak':'#417505',
                          'HuangTi':'#880E4F','Original McKenna':'#6A1B9A'}
        for _rp in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
            _pks = resonance_data[_rp]['peaks']
            _vls = resonance_data[_rp]['valleys']
            _pk_x  = [date_to_x[d] for d, _ in _pks if d in date_to_x]
            _pk_y  = [v             for d, v in _pks if d in date_to_x]
            _pk_cd = [d             for d, _ in _pks if d in date_to_x]
            _vl_x  = [date_to_x[d] for d, _ in _vls if d in date_to_x]
            _vl_y  = [v             for d, v in _vls if d in date_to_x]
            _vl_cd = [d             for d, _ in _vls if d in date_to_x]
            if _pk_x:
                fig.add_trace(go.Scatter(
                    x=_pk_x, y=_pk_y, mode='markers',
                    marker=dict(color=_peak_colors[_rp], size=7, symbol='triangle-up',
                                line=dict(color='white', width=0.5)),
                    name=f'{_rp} Peak', customdata=_pk_cd, legendgroup=f'res_{_rp}',
                    hovertemplate=f'{_rp} Peak<br>%{{customdata}}<br>%{{y:.4f}}<extra></extra>'
                ))
            if _vl_x:
                fig.add_trace(go.Scatter(
                    x=_vl_x, y=_vl_y, mode='markers',
                    marker=dict(color=_valley_colors[_rp], size=7, symbol='triangle-down',
                                line=dict(color='white', width=0.5)),
                    name=f'{_rp} Valley', customdata=_vl_cd, legendgroup=f'res_{_rp}',
                    hovertemplate=f'{_rp} Valley<br>%{{customdata}}<br>%{{y:.4f}}<extra></extra>'
                ))

    fig.update_layout(
        title=f"All Five Primitives — Level {st.session_state.levels}, Hex {st.session_state.processed_hex}",
        xaxis=x_axis_layout,
        yaxis_title="Novelty (inverted)" if invert else "Novelty",
        height=540, hovermode="x unified", showlegend=True,
        legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99)
    )
    st.plotly_chart(fig, width='stretch')

    st.subheader("Correlation Matrix")
    prims = ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]
    corr_data = [[round(float(np.corrcoef(all_y_values[p1], all_y_values[p2])[0, 1]), 4)
                  if p1 != p2 else 1.0
                  for p2 in prims] for p1 in prims]
    corr_df = pd.DataFrame(corr_data, index=prims, columns=prims)
    st.dataframe(corr_df.style.format("{:.3f}").background_gradient(cmap='RdYlGn', vmin=-1, vmax=1))

    with st.expander("📊 All Timewave Values"):
        all_data = pd.DataFrame({
            'Date': dates_str,
            'Days_Offset': np.round(x_days, 2),
            **{p: np.round(all_y_values[p], 6) for p in prims}
        })
        st.dataframe(all_data, height=400, width='stretch')
        st.download_button("📥 Download CSV", all_data.to_csv(index=False),
                           "timewave_all.csv", key="dl_comparison")

    # ── Difference overlay ────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Primitive Difference Overlay")
    st.caption("Shows where two primitives diverge most. Zero = identical behaviour at that point.")

    _diff_prims = ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]
    _dc1, _dc2 = st.columns(2)
    with _dc1:
        _diff_a = st.selectbox("Primitive A (positive)", _diff_prims,
                               index=0, key="diff_prim_a")
    with _dc2:
        _diff_b = st.selectbox("Primitive B (subtract)", _diff_prims,
                               index=1, key="diff_prim_b")

    if _diff_a == _diff_b:
        st.info("Select two different primitives to see their difference.")
    else:
        _y_diff = all_y_values[_diff_a] - all_y_values[_diff_b]
        # Normalise both source waves to [−1, 1] for fair visual comparison
        def _norm(arr):
            lo, hi = arr.min(), arr.max()
            span = hi - lo
            return (arr - lo) / span * 2 - 1 if span > 1e-12 else np.zeros_like(arr)

        _y_diff_norm = _norm(all_y_values[_diff_a]) - _norm(all_y_values[_diff_b])

        _fig_diff = go.Figure()

        # Faint reference lines for the two source primitives (normalised)
        _fig_diff.add_trace(go.Scatter(
            x=x_days, y=_norm(all_y_values[_diff_a]),
            mode='lines', name=_diff_a,
            line=dict(color=COLORS_PRIMITIVES[_diff_a], width=1, dash='dot'),
            opacity=0.4, customdata=dates_str,
            hovertemplate='%{customdata}<br>' + _diff_a + ': %{y:.4f}<extra></extra>'
        ))
        _fig_diff.add_trace(go.Scatter(
            x=x_days, y=_norm(all_y_values[_diff_b]),
            mode='lines', name=_diff_b,
            line=dict(color=COLORS_PRIMITIVES[_diff_b], width=1, dash='dot'),
            opacity=0.4, customdata=dates_str,
            hovertemplate='%{customdata}<br>' + _diff_b + ': %{y:.4f}<extra></extra>'
        ))

        # Filled difference trace
        _fig_diff.add_trace(go.Scatter(
            x=x_days, y=_y_diff_norm,
            mode='lines', name=f"{_diff_a} − {_diff_b}",
            line=dict(color='#E74C3C', width=2),
            fill='tozeroy',
            fillcolor='rgba(231,76,60,0.15)',
            customdata=dates_str,
            hovertemplate='%{customdata}<br>Diff: %{y:.4f}<extra></extra>'
        ))

        # Zero reference line
        _fig_diff.add_hline(y=0, line_dash="dash", line_color="grey", line_width=1)

        add_zero_vlines(_fig_diff, zero_date, zero_date2, x_days)
        add_event_overlays(_fig_diff, all_active_events, zero_date, x_days, _y_diff_norm)

        _fig_diff.update_layout(
            title=f"Normalised difference: {_diff_a} − {_diff_b}",
            xaxis=x_axis_layout,
            yaxis_title="Difference (normalised)",
            height=420,
            hovermode="x unified",
            legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99)
        )
        st.plotly_chart(_fig_diff, width='stretch')

        # Summary statistics
        _sc1, _sc2, _sc3, _sc4 = st.columns(4)
        _sc1.metric("Max divergence", f"{_y_diff_norm.max():.3f}")
        _sc2.metric("Min divergence", f"{_y_diff_norm.min():.3f}")
        _sc3.metric("Mean |diff|", f"{np.abs(_y_diff_norm).mean():.3f}")
        _sc4.metric("Std deviation", f"{_y_diff_norm.std():.3f}")

# ------------------------------------------------------------------
# Resonance tab (6)
# ------------------------------------------------------------------
if active_tab == "Resonance":
    st.subheader(
        f"Resonance Markers — Level {st.session_state.levels}, "
        f"Hex {st.session_state.processed_hex}: "
        f"{HEXAGRAM_NAMES[st.session_state.processed_hex - 1]}"
    )
    order_note = f" (capped from {marker_order})" if effective_order < marker_order else ""
    st.caption(
        f"Order={effective_order}{order_note} · Threshold={threshold:.2f} · "
        f"Min distance={min_dist_days}d  "
        "↳ Adjust in 🎚️ Resonance Tuning (sidebar)"
    )
    st.markdown("---")

    for prim_name in ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]:
        data    = resonance_data[prim_name]
        peaks   = data['peaks']
        valleys = data['valleys']

        st.markdown(f"### {prim_name}")
        st.markdown(f"**{len(peaks)} peaks, {len(valleys)} valleys**")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**📈 Peaks (High Novelty)**")
            if peaks:
                sp = sorted(peaks, key=lambda x: x[1], reverse=True)
                for date_s, val in sp[:10]:
                    st.markdown(f"- `{date_s}` → **{val:.4f}**")
                if len(peaks) > 10:
                    with st.expander(f"Show all {len(peaks)} peaks"):
                        for date_s, val in sp:
                            st.markdown(f"- `{date_s}` → {val:.4f}")
            else:
                st.info("No peaks — lower threshold or order.")

        with col2:
            st.markdown("**📉 Valleys (Low Novelty)**")
            if valleys:
                sv = sorted(valleys, key=lambda x: x[1])
                for date_s, val in sv[:10]:
                    st.markdown(f"- `{date_s}` → **{val:.4f}**")
                if len(valleys) > 10:
                    with st.expander(f"Show all {len(valleys)} valleys"):
                        for date_s, val in sv:
                            st.markdown(f"- `{date_s}` → {val:.4f}")
            else:
                st.info("No valleys — lower threshold or order.")

        # CSV export for this primitive
        if peaks or valleys:
            res_df = pd.DataFrame(
                [(d, v, 'peak')   for d, v in peaks] +
                [(d, v, 'valley') for d, v in valleys],
                columns=['Date', 'Novelty', 'Type']
            ).sort_values('Date')
            st.download_button(
                f"📥 {prim_name} markers CSV",
                res_df.to_csv(index=False),
                f"resonance_{prim_name}.csv",
                key=f"dl_res_{prim_name}"
            )
        st.markdown("---")

# _anim_norm defined at module level — NOT inside the button block.
# Defining it inside `if st.button():` recreates the function object
# on every click, which can cause Streamlit script-hash instability.
def _anim_norm(y):
    lo, hi = y.min(), y.max()
    span = hi - lo
    if span < 1e-12:
        return np.zeros_like(y)
    return (y - lo) / span * 2.0 - 1.0

# ------------------------------------------------------------------
# Animation tab (7)
# ------------------------------------------------------------------
if active_tab == "Animation":
    st.subheader("Wave Unfold Animation — Level by Level")
    st.caption("Change the number of Levels (1–16) in the sidebar to control the progression depth, "
               "then press ▶ Build Animation.")

    _ANIM_PRIMS = ["Kelley", "Watkins", "Sheliak", "HuangTi", "Original McKenna"]
    anim_col1, anim_col2 = st.columns([2, 1])
    with anim_col1:
        # key= alone: Streamlit reads initial value from session_state automatically.
        # DO NOT pass index= — it creates a stale-value conflict on every
        # primitive change, forcing a second rerun that resets the tab to 0.
        anim_prim = st.selectbox("Primitive", _ANIM_PRIMS, key="anim_prim")
    with anim_col2:
        anim_resolution = st.slider(
            "Resolution", 200, 1200,
            step=100, key="anim_resolution",
            help="Points per frame. Higher = more detail, slower to build."
        )

    # key= alone: no value= to avoid stale pre-click/post-click state mismatch.
    anim_show_resonance = st.checkbox(
        "Show resonance markers on animation frames",
        key="anim_show_resonance"
    )

    # ── Build button: ONLY computes and stores — never renders chart here ──
    # CRITICAL: No st.spinner() here. st.spinner() streams a UI state change
    # (spinner on → off) which disrupts tab widget state and resets to tab 0.
    # Streamlit shows a running indicator in the page title automatically.
    # The chart renders BELOW this block from session_state — no tab flip.
    if st.button("▶ Build Animation", use_container_width=True):
        n_levels = levels  # use local var — guaranteed current, not stale session state

        # Compute silently — no spinner, no streaming UI updates
        anim_x     = np.linspace(-days_before_total, days_after_total, int(anim_resolution))
        anim_dates = generate_date_labels(zero_date, anim_x)

        raw_frames  = []
        for lvl in range(1, n_levels + 1):
            y_raw = calculate_timewave(anim_prim, lvl, st.session_state.processed_hex,
                                      anim_x, invert)
            raw_frames.append((lvl, y_raw))

        norm_frames = [(lvl, _anim_norm(y)) for lvl, y in raw_frames]

        d2x = {d: anim_x[i] for i, d in enumerate(anim_dates)}
        anim_peaks_list, anim_valleys_list = [], []
        for lvl, y_norm in norm_frames:
            if anim_show_resonance:
                pk, vl = compute_resonance_markers(
                    tuple(y_norm.tolist()), tuple(anim_dates),
                    threshold, effective_order, min_dist_days,
                    tuple(anim_x.tolist())
                )
                pk_x = [d2x[d] for d, _ in pk if d in d2x]
                pk_y = [v       for d, v in pk if d in d2x]
                vl_x = [d2x[d] for d, _ in vl if d in d2x]
                vl_y = [v       for d, v in vl if d in d2x]
            else:
                pk_x, pk_y, vl_x, vl_y = [], [], [], []
            anim_peaks_list.append((pk_x, pk_y))
            anim_valleys_list.append((vl_x, vl_y))

        # Store all computed data in session_state — chart built below
        st.session_state.anim_data = {
            'norm_frames':       [(lvl, y.tolist()) for lvl, y in norm_frames],
            'anim_peaks_list':   anim_peaks_list,
            'anim_valleys_list': anim_valleys_list,
            'anim_x':            anim_x.tolist(),
            'anim_dates':        anim_dates,
            'anim_prim':         anim_prim,
            'n_levels':          n_levels,
            'processed_hex':     st.session_state.processed_hex,
            'anim_resolution':   int(anim_resolution),
            'anim_show_resonance': anim_show_resonance,
            # Snapshot events at build time so they are correct in the chart
            'all_active_events': list(all_active_events),
        }

    # ── Chart render — always runs when data exists, outside button block ──
    # This is the key: chart rendering is decoupled from the button press.
    # After Build stores data, every subsequent rerun finds the data and
    # renders the chart — no tab flip possible.
    _ad = st.session_state.get('anim_data')
    if _ad:
        # Warn if the stored animation is out of sync with current settings
        _stale_reasons = []
        if _ad['anim_prim'] != anim_prim:
            _stale_reasons.append(f"primitive changed ({_ad['anim_prim']} → {anim_prim})")
        if _ad['processed_hex'] != st.session_state.processed_hex:
            _stale_reasons.append(f"hexagram changed ({_ad['processed_hex']} → {st.session_state.processed_hex})")
        if _ad['n_levels'] != levels:
            _stale_reasons.append(f"levels changed ({_ad['n_levels']} → {levels})")
        if _stale_reasons:
            st.warning(
                "⚠️ Animation may be outdated — " + "; ".join(_stale_reasons) +
                ". Press **▶ Build Animation** to update."
            )
        norm_frames       = [(lvl, np.array(y)) for lvl, y in _ad['norm_frames']]
        anim_peaks_list   = _ad['anim_peaks_list']
        anim_valleys_list = _ad['anim_valleys_list']
        anim_x            = np.array(_ad['anim_x'])
        anim_dates        = _ad['anim_dates']
        _prim             = _ad['anim_prim']
        n_levels          = _ad['n_levels']
        _hex              = _ad['processed_hex']
        _res              = _ad['anim_resolution']
        _show_res         = _ad['anim_show_resonance']

        lvl1, y_norm1 = norm_frames[0]
        pk_x1, pk_y1  = anim_peaks_list[0]
        vl_x1, vl_y1  = anim_valleys_list[0]

        anim_fig = go.Figure()
        anim_fig.add_trace(go.Scatter(
            x=anim_x, y=y_norm1, mode='lines',
            line=dict(color=COLORS_PRIMITIVES[_prim], width=2),
            name=_prim, customdata=anim_dates,
            hovertemplate='%{customdata}<br>Novelty: %{y:.3f}<extra></extra>'
        ))
        anim_fig.add_trace(go.Scatter(
            x=pk_x1, y=pk_y1, mode='markers',
            marker=dict(color='limegreen', size=8, symbol='triangle-up',
                        line=dict(color='darkgreen', width=1)),
            name='Peaks', showlegend=_show_res, visible=True
        ))
        anim_fig.add_trace(go.Scatter(
            x=vl_x1, y=vl_y1, mode='markers',
            marker=dict(color='tomato', size=8, symbol='triangle-down',
                        line=dict(color='darkred', width=1)),
            name='Valleys', showlegend=_show_res, visible=True
        ))

        plotly_frames = []
        for i, (lvl, y_norm) in enumerate(norm_frames):
            pk_x, pk_y = anim_peaks_list[i]
            vl_x, vl_y = anim_valleys_list[i]
            plotly_frames.append(go.Frame(
                traces=[0, 1, 2],
                data=[
                    go.Scatter(
                        x=anim_x, y=y_norm, mode='lines',
                        line=dict(color=COLORS_PRIMITIVES[_prim], width=2),
                        customdata=anim_dates,
                        hovertemplate='%{customdata}<br>Novelty: %{y:.3f}<extra></extra>'
                    ),
                    go.Scatter(x=pk_x, y=pk_y, mode='markers',
                               marker=dict(color='limegreen', size=8, symbol='triangle-up')),
                    go.Scatter(x=vl_x, y=vl_y, mode='markers',
                               marker=dict(color='tomato', size=8, symbol='triangle-down')),
                ],
                name=str(lvl),
                layout=go.Layout(
                    title_text=f"{_prim} — Level {lvl} / {n_levels}",
                    yaxis=dict(range=[-1.15, 1.15], fixedrange=False)
                )
            ))

        anim_fig.frames = plotly_frames

        anim_x_layout = make_x_axis_layout(anim_x, anim_dates)
        anim_fig.update_layout(
            title=f"{_prim} — Hex {_hex}, levels 1→{n_levels}  (Y normalised per level)",
            xaxis=anim_x_layout,
            yaxis=dict(title="Novelty (normalised)", range=[-1.15, 1.15], fixedrange=False),
            height=580,
            updatemenus=[dict(
                type="buttons", showactive=False,
                y=1.15, x=0.0, xanchor="left",
                buttons=[
                    dict(label="▶ Play", method="animate",
                         args=[None, dict(frame=dict(duration=800, redraw=True),
                                         fromcurrent=True, mode="immediate")]),
                    dict(label="⏸ Pause", method="animate",
                         args=[[None], dict(frame=dict(duration=0, redraw=False),
                                            mode="immediate")])
                ]
            )],
            sliders=[dict(
                active=0,
                steps=[
                    dict(method="animate",
                         args=[[str(lvl)],
                               dict(mode="immediate",
                                    frame=dict(duration=0, redraw=True),
                                    transition=dict(duration=0))],
                         label=f"L{lvl}")
                    for lvl, _ in norm_frames
                ],
                x=0, y=0, len=1.0,
                currentvalue=dict(prefix="Level: ", visible=True, xanchor="center"),
                pad=dict(b=10, t=50),
                transition=dict(duration=0)
            )]
        )

        zero_x_val = float(anim_x[int(np.argmin(np.abs(anim_x)))])
        anim_fig.add_vline(x=zero_x_val, line_dash="dash", line_color="red",
                           line_width=2, annotation_text="Zero Date",
                           annotation_position="top right")

        # Add historical + custom event overlays (snapshot stored at build time)
        _anim_events = _ad.get('all_active_events', [])
        if _anim_events:
            _anim_y_ref = np.array([y for _, y_arr in norm_frames[:1] for y in y_arr])
            add_event_overlays(anim_fig, _anim_events, zero_date, anim_x,
                               _anim_y_ref if len(_anim_y_ref) > 0 else None)

        st.plotly_chart(anim_fig, width='stretch')
        st.caption(
            f"Primitive: **{_prim}** · Hex **{_hex}** · "
            f"Levels **1–{n_levels}** · **{_res}** pts/frame · "
            f"Y axis normalised per level"
        )
    else:
        st.info("Set Levels in the sidebar, choose a Primitive above, then press **▶ Build Animation**.")

# ------------------------------------------------------------------
# Hexagrams tab (8)
# ------------------------------------------------------------------
if active_tab == "Hexagrams":
    st.subheader("64 Hexagrams — King Wen Sequence")
    hex_cols = st.columns(8)
    for i in range(64):
        with hex_cols[i % 8]:
            with st.container(border=True):
                st.markdown(
                    f"<div style='text-align:center;font-weight:bold'>{i+1}</div>",
                    unsafe_allow_html=True
                )
                hpng = draw_hexagram(i + 1)
                st.image(hpng, use_container_width=True)
                short = HEXAGRAM_NAMES[i].split('.')[1].strip()
                st.markdown(
                    f"<div style='text-align:center;font-size:0.72em'>{short}</div>",
                    unsafe_allow_html=True
                )

# ------------------------------------------------------------------
# Manual tab (9)
# ------------------------------------------------------------------
if active_tab == "Manual":
    st.subheader("📖 Manual")

    _MANUAL_TABS = ["Theory", "The Primitives", "How to..", "Controls", "Interpretations", "Quick Troubleshooting"]
    manual_active_tab = st.radio(
        "##manualnav", _MANUAL_TABS, horizontal=True,
        key="manual_active_tab", label_visibility="collapsed"
    )

    if manual_active_tab == "Theory":
        st.markdown("""
Theory created by Terence McKenna (1946–2000) and mathematically formalized by Peter Meyer.

**Core idea:**
History is **not linear** — it follows a **fractal wave of novelty** derived from the King Wen sequence of the I Ching.
Novelty = degree of newness / interconnectedness / complexity
Habit = repetition / stasis

The wave supposedly increases in novelty (decreases in habit) as we approach a theoretical singularity / eschaton / zero point (most famously proposed as **21 December 2012**).

---

### Key Literature & Primary Sources

**Essential reading / viewing (chronological)**

- Terence McKenna — **True Hallucinations** (1993) → first public mention of the timewave idea
- Terence McKenna — **The Archaic Revival** (1991), essay "Timewave Zero"
- Peter Meyer — **The Timewave Zero software** (DOS versions 1990s) — original implementation
  http://www.fractal-timewave.com/ (archived mirrors still exist 2026)

**Modern / explanatory works**

- Matthew Watkins — "The Mathematics of Timewave Zero" (late 1990s – early 2000s essays)
  → https://www.maths.surrey.ac.uk/hosted-sites/timewave/ (sometimes still online)
- Dennis McKenna — podcast & interview comments (2010s–2020s)
  frequently says the 2012 date was **symbolic**, not literal
- **Lunar calendar coincidence paper** by various authors (~2010–2015)
  shows how closely the wave matches certain Chinese calendrical cycles

**Video introductions** (still watchable 2026)

- "Timewave Zero explained" — Lorenzo Hagerty / Psychedelic Salon (~45 min)
- Several Terence McKenna Esalen lectures (1980s–90s) that mention it
- "novelty theory – 15 minute summary" various YouTube channels (search 2024–2025 uploads)

---

*Enjoy exploring — whether as serious theory, beautiful math toy, or cultural artifact.*
        """)

    if manual_active_tab == "The Primitives":
        st.markdown("""
### The Primitives

| Primitive | Origin / Author | Character | Typical use today |
|---|---|---|---|
| **Kelley** | Dennis Kelley | Most commonly used 1980s–2000s | Default / historical benchmark |
| **Watkins** | Matthew Watkins | Slightly different scaling | Often compared with Kelley |
| **Sheliak** | Sheliak (pseudonym) | Very low amplitude early levels | Sometimes shows sharper late peaks |
| **HuangTi** | HuangTi (pseudonym) | Minimalist / reduced values | Cleanest-looking small-level plots |
| **Original McKenna** | Raw first differences | No secondary processing | Closest to McKenna's early sketches |

Most serious users look mainly at **Kelley** and **Watkins**, sometimes **Original McKenna**.
        """)

    if manual_active_tab == "How to..":
        st.markdown("""
### How to use this app

1. **Choose zero date**
   Most common: `2012-12-21` (CE)
   Other popular: Mayan start `BCE 3114-08-11`, personal significant dates

2. **Select hexagram** (1–64)
   - Use **Throw** 🎲 for random/I Ching-style cast
   - Or manually choose and press **Process**

3. **Set levels** (usually 5–9)
   - 6–7 levels → good balance of detail vs. readability
   - >13 levels → numbers become huge (clamped automatically)

4. **Look at plots**
   - Single primitive tabs = clean view
   - **Comparison** tab = all five at once + correlation matrix

5. **Resonance markers** (optional)
   Peaks (green) = novelty surges / breakthroughs
   Valleys (red) = habit / consolidation periods

6. **Animation tab** — most visually impressive way to understand fractality
        """)

    if manual_active_tab == "Controls":
        st.markdown("""
### Most Important Controls (sidebar)

| Setting | Typical / recommended value | Effect |
|:---|:---|:---|
| **Levels** | 6–8 | Fractal depth — more = finer structure |
| **Invert novelty** | ☑ checked | High values = high novelty (McKenna style) |
| **Resolution** | 1000–1600 | Plot smoothness vs loading time |
| **Resonance threshold** | 0.04–0.08 | How sharp a peak/valley must be |
| **Extrema order** | 4–12 | How local the extremum must be |
        """)

    if manual_active_tab == "Interpretations":
        st.markdown("""
### Common Interpretations in 2025–2026

| Interpretation | Roughly when popular | Current status (2026) |
|---|---|---|
| Literal end-of-history Dec 2012 | 1995–2013 | Almost nobody defends literally anymore |
| Symbolic attractor / phase transition | 2013–2020 | Still the mainstream "believer" position |
| Interesting fractal toy / art project | 2018–2026 | Most scientifically-minded users today |
| Calendrical / cultural pattern detector | ~2022–present | Growing academic fringe interest |
        """)

    if manual_active_tab == "Quick Troubleshooting":
        st.markdown("""
### Quick Troubleshooting

- **Nothing happens after Throw / Process** → press Process again (sometimes Streamlit state glitch)
- **No resonance markers** → lower threshold to ~0.03 or reduce order to 3–5
- **Graph looks flat** → increase levels or check Invert checkbox
- **Overflow warning** → reduce levels (especially >13)
        """)

# ------------------------------------------------------------------
# Math tab (10)
# ------------------------------------------------------------------
if active_tab == "Math":
    st.subheader("📐 Mathematical Foundations of Timewave Zero")
    st.markdown(
        "This tab explains the mathematics behind the timewave in four ways — "
        "choose whichever suits you best."
    )

    _MATH_TABS = ["Equations", "Live Calculator", "Python Code", "Excel Export"]
    math_active_tab = st.radio(
        "##mathnav", _MATH_TABS, horizontal=True,
        key="math_active_tab", label_visibility="collapsed"
    )

    # ── A: Proper typeset equations ─────────────────────────────────────────
    if math_active_tab == "Equations":
        st.markdown("### 1 · The Primitive Sequence — 384 values")
        st.markdown(
            "The entire wave is built from a single sequence of **384 numbers** called the "
            "**primitive** P. It encodes the pattern of change between consecutive I Ching hexagrams."
        )
        st.markdown("**Step 1** — Represent each hexagram as a 6-bit vector: solid line = 1, broken = 0.")
        st.markdown("**Step 2** — Compute the Hamming distance between consecutive hexagrams:")
        st.latex(
            r"d_n = \sum_{i=1}^{6} \left| H_n^{(i)} - H_{n+1}^{(i)} \right|"
            r"\qquad n = 1,\ldots,64 \quad (\text{circular: } H_{65} = H_1)"
        )
        st.markdown(
            "**Step 3** — Reverse the 64 distances, tile 6 times → "
            "primitive **P = [p₀, p₁, …, p₃₈₃]**."
        )
        st.markdown("Other primitives (Watkins, Sheliak, HuangTi, Original McKenna) use variations "
                    "of this derivation with different scaling or reversal rules.")

        st.markdown("---")
        st.markdown("### 2 · The Core Wave Equation")
        st.markdown(
            "The novelty N at time **t** (days from the zero date) "
            "across **ℓ** fractal levels is:"
        )
        st.latex(
            r"N(t,\,\ell) \;=\; \sum_{k=0}^{\ell-1} 64^{\,k} \cdot \mathrm{lerp}\!\left(P,\; \frac{t}{64^k}\right)"
        )
        st.markdown("where **lerp** linearly interpolates between two adjacent primitive values:")
        st.latex(
            r"\mathrm{lerp}(P,\,x) \;=\; "
            r"P_{\lfloor x \rfloor \bmod 384}"
            r"\;+\; \bigl(x - \lfloor x \rfloor\bigr)"
            r"\cdot\Bigl("
            r"P_{(\lfloor x \rfloor+1)\bmod 384}"
            r"- P_{\lfloor x \rfloor \bmod 384}"
            r"\Bigr)"
        )
        st.markdown("Expanding the full sum:")
        st.latex(
            r"N(t,\ell) = \sum_{k=0}^{\ell-1} 64^k \left["
            r"P_{\left\lfloor t/64^k \right\rfloor \bmod 384}"
            r"+ \left\{t/64^k\right\}"
            r"\Bigl(P_{\left(\lfloor t/64^k \rfloor+1\right)\bmod 384}"
            r"- P_{\lfloor t/64^k \rfloor \bmod 384}\Bigr)"
            r"\right]"
        )
        st.caption("{x} denotes the fractional part of x.")

        st.markdown("---")
        st.markdown("### 3 · Hexagram Offset")
        st.markdown(
            "Hexagram **h** rotates the starting position within the 384-value cycle "
            "before computation:"
        )
        st.latex(
            r"P^{(h)}_i \;=\; P_{\,(i\;+\;(h-1)\times 6)\;\bmod\;384}"
        )

        st.markdown("---")
        st.markdown("### 4 · Inversion Convention")
        st.markdown(
            "McKenna inverted the wave: high novelty = valleys (breakthroughs), "
            "low novelty = peaks (habit)."
        )
        st.latex(r"y(t) \;=\; -\,N(t,\,\ell)")
        st.markdown("The **Invert** toggle in the sidebar applies this sign flip.")

        st.markdown("---")
        st.markdown("### 5 · Resonance Peaks and Valleys")
        st.markdown("A point tₘ is a **peak** if it exceeds all neighbours within a window:")
        st.latex(
            r"y(t_m) > y(t_{m\pm j}) \quad \text{for all } j=1,\ldots,\mathrm{order}"
        )
        st.markdown(
            "Points are further filtered by a sharpness threshold on the normalised "
            "derivative |Δy/Δt| and a minimum temporal separation."
        )

        st.markdown("---")
        st.markdown("### 6 · Complete Formula")
        st.latex(
            r"y(\delta) \;=\; -\sum_{k=0}^{\ell-1} 64^{\,k} \cdot \mathrm{lerp}"
            r"\!\left(P^{(h)},\;\frac{\delta}{64^k}\right)"
        )
        st.markdown(
            "δ = days from zero date · ℓ = levels · h = hexagram · "
            "P⁽ʰ⁾ = hexagram-rotated primitive."
        )

    # ── B: Live interactive calculator ──────────────────────────────────────
    if math_active_tab == "Live Calculator":
        st.markdown("### 🧮 Live Novelty Calculator")
        st.markdown(
            "Enter a time offset and see the computation broken down level by level. "
            "Uses the same primitives and hexagram data as the main plots."
        )

        _c1, _c2, _c3 = st.columns(3)
        with _c1:
            calc_prim = st.selectbox("Primitive", list(PRIMITIVES.keys()), key="calc_prim")
        with _c2:
            calc_hex = st.number_input("Hexagram (1–64)", 1, 64,
                                       step=1, key="calc_hex")
        with _c3:
            calc_levels = st.number_input("Levels (max 8)", 1, 8,
                                          step=1, key="calc_levels",
                                          help="Capped at 8 to avoid float overflow in table display")

        calc_t = st.number_input(
            "Time offset t (days from zero date — negative = before)",
            step=1.0, format="%.1f", key="calc_t"
        )

        # Compute step by step
        _prim_calc = PRIMITIVES[calc_prim].copy()
        _prim_calc = np.roll(_prim_calc, -(int(calc_hex) - 1) * 6)

        _rows, _total, _scale, _overflow = [], 0.0, 1.0, False
        for _k in range(int(calc_levels)):
            _x    = float(calc_t) / _scale if _scale != 0 else 0.0
            _n    = int(np.floor(_x)) % 384
            _frac = _x - np.floor(_x)
            _pn   = float(_prim_calc[_n])
            _pn1  = float(_prim_calc[(_n + 1) % 384])
            _lerp = _pn + _frac * (_pn1 - _pn)
            _con  = _scale * _lerp
            if not np.isfinite(_con) or abs(_con) > 1e15:
                _overflow = True
                _rows.append({"k": _k, "64ᵏ": f"{_scale:.2e}", "t/64ᵏ": f"{_x:.4f}",
                              "n=⌊·⌋ mod 384": _n, "P[n]": f"{_pn:.3f}",
                              "P[n+1]": f"{_pn1:.3f}", "frac": f"{_frac:.4f}",
                              "lerp": f"{_lerp:.3f}", "64ᵏ·lerp": "OVERFLOW"})
            else:
                _total += _con
                _rows.append({"k": _k, "64ᵏ": f"{_scale:.2e}", "t/64ᵏ": f"{_x:.4f}",
                              "n=⌊·⌋ mod 384": _n, "P[n]": f"{_pn:.3f}",
                              "P[n+1]": f"{_pn1:.3f}", "frac": f"{_frac:.4f}",
                              "lerp": f"{_lerp:.3f}", "64ᵏ·lerp": f"{_con:.4e}"})
            _scale *= 64.0

        st.markdown(f"#### Breakdown at t = **{calc_t}** days")
        st.dataframe(pd.DataFrame(_rows), use_container_width=True)

        if not _overflow:
            _m1, _m2 = st.columns(2)
            with _m1:
                st.metric("N(t, ℓ)  — raw sum", f"{_total:.6e}")
            with _m2:
                st.metric("y(t) = −N  — displayed value", f"{-_total:.6e}")
            st.caption(
                f"Primitive **{calc_prim}** · Hex **{calc_hex}** · "
                f"Levels **{calc_levels}** · t = **{calc_t}** days"
            )
        else:
            st.warning("⚠️ Overflow at this level — reduce Levels or use the main plot.")

        with st.expander("📖 Column guide"):
            st.markdown(
                "| Column | Meaning |\n"
                "|--------|---------|\n"
                "| **k** | Fractal level (0 = finest) |\n"
                "| **64ᵏ** | Scale factor — grows as 1, 64, 4096, … |\n"
                "| **t/64ᵏ** | Time scaled to this level |\n"
                "| **n** | Integer index into P (0–383) |\n"
                "| **P[n], P[n+1]** | Primitive values being interpolated |\n"
                "| **frac** | Fractional position between P[n] and P[n+1] |\n"
                "| **lerp** | Interpolated primitive value |\n"
                "| **64ᵏ·lerp** | This level's contribution to N(t) |"
            )

    # ── C: Python code snippet ───────────────────────────────────────────────
    if math_active_tab == "Python Code":
        st.markdown("### 🐍 Self-contained Python Implementation")
        st.markdown(
            "No dependencies except NumPy. Copy-paste into any Python environment "
            "to reproduce the exact computation used in this app."
        )
        _py_code = (
            "import numpy as np\n\n"
            "# Kelley primitive — 384 values (the full array used by the app).\n"
            "# Swap for any other primitive from the app source.\n"
            "KELLEY_384 = [\n"
            "    50,48,48,40,41,38,31,31,31,22,25,28,25,59,44,62,54,48,38,22,20,20,22,23,\n"
            "    27,36,40,48,72,62,63,52,49,41,43,40,39,42,44,46,42,41,38,37,37,39,37,41,\n"
            "    45,41,44,44,42,47,59,69,69,60,63,60,42,42,37,40,49,47,41,37,44,34,29,29,\n"
            "    43,54,49,37,33,43,46,39,37,42,47,43,37,37,39,42,40,41,35,38,40,41,37,38,\n"
            "    38,33,35,45,50,55,35,29,35,37,39,38,43,52,54,65,31,41,44,44,42,29,24,24,\n"
            "    32,39,32,38,44,34,39,46,51,64,60,64,49,46,51,48,48,45,48,49,51,51,34,29,\n"
            "    41,41,46,44,46,50,45,46,44,45,43,44,58,40,44,40,43,42,47,32,31,30,32,26,\n"
            "    62,79,75,71,79,76,31,31,33,38,41,41,33,18,23,27,29,30,38,47,47,30,31,28,\n"
            "    34,28,24,38,37,34,33,35,29,32,27,26,27,33,18,36,28,22,22,26,24,24,26,27,\n"
            "    35,40,44,52,76,66,47,28,29,27,33,36,43,40,32,44,32,31,36,17,13,37,39,43,\n"
            "    39,31,28,38,36,41,53,63,63,54,57,54,42,42,37,40,49,47,51,43,36,46,53,53,\n"
            "    51,48,51,53,51,51,44,39,41,34,25,29,49,61,61,34,32,33,41,26,26,29,27,28,\n"
            "    44,43,45,41,44,49,47,43,35,53,31,36,37,36,34,31,33,29,22,22,20,49,42,46,\n"
            "    66,65,74,80,86,76,55,58,53,56,60,56,31,34,29,42,30,39,30,57,59,75,58,67,\n"
            "    41,41,38,32,34,38,45,46,44,45,43,44,76,70,74,34,35,34,39,38,37,40,38,40,\n"
            "    32,27,29,47,49,48,37,37,39,44,47,47,45,54,59,49,49,42,30,27,27,30,25,28\n"
            "]\n"
            "P = np.array(KELLEY_384, dtype=float)  # 384 values — do NOT tile\n\n\n"
            "def novelty(t, levels, primitive, hexagram=1, invert=True):\n"
            '    """\n'
            "    Compute timewave novelty at time t (days from zero date).\n\n"
            "    Parameters\n"
            "    ----------\n"
            "    t         : float — days from zero date (negative = before)\n"
            "    levels    : int   — fractal depth (1-16)\n"
            "    primitive : array — 384-value primitive P\n"
            "    hexagram  : int   — hexagram 1-64 (rotates primitive)\n"
            "    invert    : bool  — negate result (McKenna convention)\n"
            '    """\n'
            "    P = np.roll(primitive.copy(), -(hexagram - 1) * 6)\n"
            "    total, scale = 0.0, 1.0\n"
            "    for k in range(levels):\n"
            "        x    = t / scale\n"
            "        n    = int(np.floor(x)) % 384\n"
            "        frac = x - np.floor(x)\n"
            "        lerp = P[n] + frac * (P[(n + 1) % 384] - P[n])\n"
            "        total += scale * lerp\n"
            "        scale *= 64.0\n"
            "    return -total if invert else total\n\n\n"
            "def novelty_array(t_array, levels, primitive, hexagram=1, invert=True):\n"
            '    """Vectorised version for NumPy arrays of t values."""\n'
            "    P = np.roll(primitive.copy(), -(hexagram - 1) * 6)\n"
            "    y, scale = np.zeros(len(t_array), dtype=float), 1.0\n"
            "    for k in range(levels):\n"
            "        x    = t_array / scale\n"
            "        n    = np.floor(x).astype(int) % 384\n"
            "        frac = x - np.floor(x)\n"
            "        lerp = P[n] + frac * (P[(n + 1) % 384] - P[n])\n"
            "        y   += scale * lerp\n"
            "        scale *= 64.0\n"
            "    return -y if invert else y\n\n\n"
            "# Example\n"
            'if __name__ == "__main__":\n'
            "    val = novelty(-365.0, levels=6, primitive=P, hexagram=1)\n"
            '    print(f"Novelty at t=-365 days: {val:.6e}")\n\n'
            "    t = np.linspace(-365*25, 365*5, 1000)\n"
            "    y = novelty_array(t, levels=6, primitive=P, hexagram=1)\n"
            '    print(f"Peak at t = {t[y.argmax()]:.1f} days from zero date")\n'
        )
        st.code(_py_code, language="python")
        st.download_button(
            "📥 Download timewave.py", _py_code,
            file_name="timewave.py", mime="text/x-python",
            key="dl_python_snippet"
        )

    # ── D: Excel export ──────────────────────────────────────────────────────
    if math_active_tab == "Excel Export":
        st.markdown("### 📊 Excel Workbook Export")
        st.markdown(
            "Generates a downloadable `.xlsx` file with three sheets: "
            "the primitive array, the computed wave values, and a formula guide. "
            "No macros — just numbers and explanations."
        )
        st.info(
            "⚠️ Works well for Levels 1–5. At Level 6+ numbers exceed Excel's "
            "float range and cells show #NUM!. Use the Python code for higher levels."
        )

        _xc1, _xc2, _xc3 = st.columns(3)
        with _xc1:
            xl_prim = st.selectbox("Primitive", list(PRIMITIVES.keys()), key="xl_prim")
        with _xc2:
            xl_hex = st.number_input("Hexagram (1–64)", 1, 64,
                                     step=1, key="xl_hex")
        with _xc3:
            xl_levels = st.number_input("Levels (max 5)", 1, 5,
                                        step=1, key="xl_levels")

        xl_points     = st.slider("Time points", 50, 500, step=50, key="xl_points")
        _xd1, _xd2   = st.columns(2)
        with _xd1:
            xl_before = st.number_input("Days before zero date", 1, 5000000,
                                        step=365, key="xl_before")
        with _xd2:
            xl_after  = st.number_input("Days after zero date", 1, 2000000,
                                        step=365, key="xl_after")

        if st.button("⚙️ Generate Excel workbook", key="gen_excel"):
            import io
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                wb  = openpyxl.Workbook()
                _hf = PatternFill("solid", fgColor="1565C0")
                _hfont = Font(color="FFFFFF", bold=True)

                # Sheet 1 — primitive
                ws_p = wb.active
                ws_p.title = "Primitive"
                _pr = np.roll(PRIMITIVES[xl_prim].copy(), -(int(xl_hex) - 1) * 6)
                ws_p["A1"] = f"Primitive: {xl_prim}  |  Hexagram offset: {xl_hex}"
                ws_p["A1"].font = Font(bold=True, size=12)
                for col, hdr in enumerate(["Index (0–383)", "P[index]"], 1):
                    c = ws_p.cell(row=2, column=col, value=hdr)
                    c.fill, c.font = _hf, _hfont
                for i, v in enumerate(_pr):
                    ws_p.cell(row=i+3, column=1, value=i)
                    ws_p.cell(row=i+3, column=2, value=float(v))
                ws_p.column_dimensions["A"].width = 18
                ws_p.column_dimensions["B"].width = 14

                # Sheet 2 — wave values
                ws_w = wb.create_sheet("Wave")
                _t   = np.linspace(-xl_before, xl_after, int(xl_points))
                _hdrs = ["t (days)", "N(t) raw", "y = −N"] + [f"Level {k}" for k in range(int(xl_levels))]
                ws_w["A1"] = f"{xl_prim}  |  Hex {xl_hex}  |  Levels {xl_levels}  |  {xl_points} points"
                ws_w["A1"].font = Font(bold=True, size=11)
                ws_w.merge_cells(f"A1:{get_column_letter(len(_hdrs))}1")
                for ci, h in enumerate(_hdrs, 1):
                    c = ws_w.cell(row=2, column=ci, value=h)
                    c.fill, c.font = _hf, _hfont
                    c.alignment = Alignment(horizontal="center")

                _sc, _lcs = 1.0, []
                for _k in range(int(xl_levels)):
                    _x   = _t / _sc
                    _n   = np.floor(_x).astype(int) % 384
                    _fr  = _x - np.floor(_x)
                    _lr  = _pr[_n] + _fr * (_pr[(_n+1)%384] - _pr[_n])
                    _lcs.append(_sc * _lr)
                    _sc *= 64.0
                _N = sum(_lcs)
                _Y = -_N

                def _safe(v):
                    return round(float(v), 6) if np.isfinite(v) and abs(v) < 1e15 else "OVERFLOW"

                for ri, tv in enumerate(_t):
                    r = ri + 3
                    ws_w.cell(row=r, column=1, value=round(float(tv), 3))
                    ws_w.cell(row=r, column=2, value=_safe(_N[ri]))
                    ws_w.cell(row=r, column=3, value=_safe(_Y[ri]))
                    for ki, lc in enumerate(_lcs):
                        ws_w.cell(row=r, column=4+ki, value=_safe(lc[ri]))

                for ci in range(1, len(_hdrs)+1):
                    ws_w.column_dimensions[get_column_letter(ci)].width = 16

                # Sheet 3 — formula guide
                ws_f = wb.create_sheet("Formula Guide")
                ws_f["A1"] = "How the Timewave is Computed"
                ws_f["A1"].font = Font(bold=True, size=13)
                _guide = [
                    "", "For each time point t and each level k (0 to levels-1):",
                    "  x  =  t / 64^k                  (scale time to this level)",
                    "  n  =  FLOOR(x, 1) MOD 384        (index into primitive array)",
                    "  f  =  x - FLOOR(x, 1)            (fractional part of x)",
                    "  lerp  =  P[n] + f * (P[n+1] - P[n])   (linear interpolation)",
                    "  contrib_k  =  64^k * lerp         (this level's contribution)",
                    "",
                    "Total:    N(t) = SUM of all contrib_k",
                    "Display:  y(t) = -N(t)              (McKenna inversion convention)",
                    "",
                    "The 'Wave' sheet contains pre-computed values for each time point.",
                    "The 'Primitive' sheet contains the rotated primitive array P.",
                    "",
                    "Note: Excel formulas for this computation would require VLOOKUP",
                    "across 384 rows for every cell — impractical for large tables.",
                    "The Python snippet (in the app's Math tab) is recommended instead.",
                ]
                for ri, line in enumerate(_guide, 2):
                    ws_f.cell(row=ri, column=1, value=line)
                ws_f.column_dimensions["A"].width = 70

                _buf = io.BytesIO()
                wb.save(_buf)
                _buf.seek(0)
                st.success(f"✅ Ready — {xl_prim}, Hex {xl_hex}, {xl_levels} levels, {xl_points} points")
                st.download_button(
                    "📥 Download timewave.xlsx", _buf,
                    file_name=f"timewave_{xl_prim}_hex{xl_hex}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_file"
                )
            except ImportError:
                st.error("❌ openpyxl not installed. Run: pip install openpyxl")

st.markdown("---")
st.caption("Timewave Zero · Terence McKenna & Peter Meyer · version 1_2026 · Streamlit + Plotly")